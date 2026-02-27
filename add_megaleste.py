import openpyxl

caminho = "planilha_modelo.xlsx"
wb = openpyxl.load_workbook(caminho)
ws = wb.active

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

proxima = ws.max_row + 1
fill = PatternFill("solid", fgColor="EEF4FF" if proxima % 2 == 0 else "FFFFFF")
border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

valores = ["megaleste.com.br", "seu_email@email.com", "sua_senha"]
for col, val in enumerate(valores, start=1):
    cell = ws.cell(row=proxima, column=col, value=val)
    cell.fill = fill
    cell.font = Font(size=11)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = border

ws.row_dimensions[proxima].height = 22
wb.save(caminho)
print(f"megaleste.com.br adicionado na linha {proxima} da planilha_modelo.xlsx")
