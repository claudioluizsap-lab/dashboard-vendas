import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Criar workbook
wb = Workbook()
ws = wb.active
ws.title = "Estoque"

# Cabeçalho
headers = ['Código', 'Descrição', 'Quantidade', 'Valor Unitário']
ws.append(headers)

# Estilizar cabeçalho
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Dados de exemplo variados
dados = [
    ['P001', 'Notebook Dell Inspiron 15', 15, 3500.00],
    ['P002', 'Mouse Logitech MX Master', 150, 45.00],
    ['P003', 'Teclado Mecânico Redragon', 80, 250.00],
    ['P004', 'Monitor LG 24" Full HD', 25, 890.00],
    ['P005', 'Webcam HD Logitech C920', 40, 180.00],
    ['P006', 'Headset Gamer HyperX', 30, 320.00],
    ['P007', 'Pen Drive 32GB SanDisk', 200, 25.00],
    ['P008', 'HD Externo 1TB Seagate', 35, 380.00],
    ['P009', 'Impressora HP LaserJet', 12, 1200.00],
    ['P010', 'Roteador WiFi TP-Link', 45, 220.00],
    ['P011', 'Switch 8 portas Intelbras', 20, 150.00],
    ['P012', 'Cabo HDMI 2m', 180, 15.00],
    ['P013', 'Mousepad Gamer Grande', 120, 20.00],
    ['P014', 'Cadeira Gamer Ergonômica', 8, 1800.00],
    ['P015', 'Mesa para Computador', 10, 650.00],
    ['P016', 'Luminária LED de Mesa', 50, 85.00],
    ['P017', 'Suporte para Monitor', 35, 95.00],
    ['P018', 'Estabilizador 500VA', 25, 180.00],
    ['P019', 'No-break 600VA', 15, 420.00],
    ['P020', 'Gabinete ATX Cooler Master', 18, 280.00],
]

# Adicionar dados
for row in dados:
    ws.append(row)

# Ajustar largura das colunas
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 15

# Salvar
wb.save('curva-abc-app/EXEMPLO-PLANILHA-MODELO.xlsx')
print('✅ Planilha modelo criada: EXEMPLO-PLANILHA-MODELO.xlsx')

# Também criar versão com dados diferentes para teste
wb2 = Workbook()
ws2 = wb2.active
ws2.title = "Produtos"

ws2.append(headers)
for cell in ws2[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Exemplo com produtos diferentes
dados2 = [
    ['001', 'Smartphone Samsung Galaxy', 50, 2500.00],
    ['002', 'Fone Bluetooth JBL', 100, 150.00],
    ['003', 'Carregador USB-C', 200, 35.00],
    ['004', 'Película de Vidro', 300, 15.00],
    ['005', 'Capa de Silicone', 250, 25.00],
    ['006', 'Tablet Samsung 10"', 30, 1800.00],
    ['007', 'Smart TV 55" LG', 15, 3200.00],
    ['008', 'Soundbar Samsung', 20, 850.00],
    ['009', 'Controle Remoto Universal', 80, 45.00],
    ['010', 'Cabo HDMI Premium 3m', 150, 35.00],
]

for row in dados2:
    ws2.append(row)

ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 35
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 15

wb2.save('curva-abc-app/EXEMPLO-ELETRONICOS.xlsx')
print('✅ Planilha de eletrônicos criada: EXEMPLO-ELETRONICOS.xlsx')

print('\n📊 Total de exemplos criados: 2 planilhas')
