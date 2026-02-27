import pandas as pd
import openpyxl

print('='*70)
print('ANÁLISE COMPLETA DA PLANILHA - planilha-teste-estoque.xlsx')
print('='*70)

# Carregar com openpyxl para ver estrutura
wb = openpyxl.load_workbook('planilha-teste-estoque.xlsx', data_only=True)
print('\n📄 SHEETS:')
for name in wb.sheetnames:
    ws = wb[name]
    print(f'   [{name}] rows={ws.max_row} cols={ws.max_column}')

# Carregar primeira linha para ver colunas
ws = wb.active
print('\n🔑 COLUNAS (primeira linha):')
for cell in ws[1]:
    print(f'   {cell.coordinate}: "{cell.value}" (tipo: {type(cell.value).__name__})')

print('\n📦 PRIMEIRAS 3 LINHAS DE DADOS:')
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=4, values_only=True), start=2):
    print(f'\n   Linha {i}:')
    for j, val in enumerate(row, start=1):
        col_name = ws.cell(1, j).value
        print(f'      {col_name}: {val} (tipo: {type(val).__name__})')

wb.close()

# Carregar com pandas
print('\n' + '='*70)
print('ANÁLISE COM PANDAS')
print('='*70)

df = pd.read_excel('planilha-teste-estoque.xlsx')
print('\n📊 INFO:')
print(df.info())

print('\n📈 PRIMEIROS 5 REGISTROS:')
print(df.head().to_string())

print('\n💰 ESTATÍSTICAS:')
print(df.describe())

print('\n🔍 VALORES ÚNICOS:')
for col in df.columns:
    print(f'   {col}: {df[col].nunique()} valores únicos')

print('\n⚠️  VALORES NULOS:')
nulls = df.isnull().sum()
if nulls.sum() == 0:
    print('   ✅ Nenhum valor nulo!')
else:
    print(nulls[nulls > 0])

# Testar cálculo da Curva ABC
print('\n' + '='*70)
print('TESTE DO CÁLCULO DA CURVA ABC')
print('='*70)

df['Valor Total'] = df['Quantidade'] * df['Valor Unitário']
df = df.sort_values('Valor Total', ascending=False)

valor_total_geral = df['Valor Total'].sum()
df['% Valor'] = (df['Valor Total'] / valor_total_geral) * 100
df['% Acumulado'] = df['% Valor'].cumsum()

df['Classe'] = df['% Acumulado'].apply(
    lambda x: 'A' if x <= 80 else ('B' if x <= 95 else 'C')
)

print(f'\n💰 Valor total: R$ {valor_total_geral:,.2f}')
print('\n📊 Distribuição:')
for classe in ['A', 'B', 'C']:
    qtd = len(df[df['Classe'] == classe])
    valor = df[df['Classe'] == classe]['Valor Total'].sum()
    perc = (valor / valor_total_geral) * 100
    print(f'   Classe {classe}: {qtd} produtos ({perc:.1f}% do valor)')

print('\n🏆 TOP 10:')
print(df[['Código', 'Descrição', 'Quantidade', 'Valor Total', '% Acumulado', 'Classe']].head(10).to_string())

print('\n✅ ANÁLISE CONCLUÍDA!')
print('='*70)
