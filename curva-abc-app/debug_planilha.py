import openpyxl
import pandas as pd

# Verificar com openpyxl (leitura direta das células)
print("=== VERIFICAÇÃO COM OPENPYXL ===")
wb = openpyxl.load_workbook('planilha-teste-estoque.xlsx')
ws = wb.active

print("\n=== CABEÇALHOS (linha 1) ===")
for i in range(1, 10):
    cell = ws.cell(1, i)
    if cell.value:
        print(f"Coluna {i}: '{cell.value}' (tipo: {type(cell.value).__name__})")
        # Mostrar código ASCII de cada caractere
        chars = [f"{c}({ord(c)})" for c in str(cell.value)]
        print(f"  Caracteres: {' '.join(chars)}")

print("\n=== PRIMEIRA LINHA DE DADOS (linha 2) ===")
for i in range(1, 10):
    cell = ws.cell(2, i)
    if cell.value is not None:
        print(f"Coluna {i}: '{cell.value}' (tipo: {type(cell.value).__name__})")

# Verificar com pandas
print("\n=== VERIFICAÇÃO COM PANDAS ===")
df = pd.read_excel('planilha-teste-estoque.xlsx')
print("\nColunas:", df.columns.tolist())
print("\nPrimeira linha:")
print(df.iloc[0])

# Verificar se há espaços ou caracteres especiais
print("\n=== ANÁLISE DE NOMES DE COLUNAS ===")
for col in df.columns:
    print(f"Coluna: '{col}'")
    print(f"  Tamanho: {len(col)} caracteres")
    print(f"  Repr: {repr(col)}")
    chars = [f"{c}({ord(c)})" for c in col]
    print(f"  Caracteres: {' '.join(chars)}")
    print()
