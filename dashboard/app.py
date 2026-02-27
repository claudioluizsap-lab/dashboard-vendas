from flask import Flask, render_template, request, Response
from waitress import serve
import pandas as pd
import numpy as np
import json, os
from datetime import datetime
from werkzeug.utils import secure_filename

app = Flask(__name__)

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

DEFAULT_EXCEL = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'Bd_dadosprojecao.xlsx'))
ALLOWED_EXT   = {'.xlsx', '.xlsm', '.xls'}

_excel_ativo  = {'path': DEFAULT_EXCEL, 'nome': 'Bd_dadosprojecao.xlsx'}

MESES_PT  = ['Janeiro','Fevereiro','Marco','Abril','Maio','Junho',
             'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
MESES_ABR = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']
DIAS_MES  = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
FATOR_SAZ = [1.05, 0.95, 1.00, 0.98, 1.02, 0.97, 0.96, 1.00, 1.03, 1.05, 1.08, 1.10]


def jsend(data, status=200):
    def fix(o):
        if isinstance(o, (np.integer,)):  return int(o)
        if isinstance(o, (np.floating,)): return float(o)
        if isinstance(o, np.ndarray):     return o.tolist()
        if isinstance(o, pd.Timestamp):   return str(o)
        raise TypeError(type(o))
    return Response(json.dumps(data, default=fix, ensure_ascii=False),
                    status=status, mimetype='application/json')


def br(v):
    return 'R$ ' + f'{float(v):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')


def load_data():
    path = _excel_ativo['path']
    if not os.path.exists(path):
        raise FileNotFoundError(f'Arquivo nao encontrado: {path}')

    xf    = pd.ExcelFile(path)
    sheet = 'Planilha1' if 'Planilha1' in xf.sheet_names else xf.sheet_names[0]
    df    = pd.read_excel(path, sheet_name=sheet)

    # detecta colunas DATA e FATURADO (insensivel a maiusculas e espacos)
    col_map  = {str(c).upper().strip(): c for c in df.columns}
    col_data = col_map.get('DATA', df.columns[0])
    col_fat  = col_map.get('FATURADO', df.columns[1])

    df = df.rename(columns={col_data: 'DATA', col_fat: 'FATURADO'})
    df['DATA']     = pd.to_datetime(df['DATA'], dayfirst=True, errors='coerce')
    df['FATURADO'] = pd.to_numeric(df['FATURADO'], errors='coerce').fillna(0.0)
    df = df.dropna(subset=['DATA']).sort_values('DATA').reset_index(drop=True)
    df['MES'] = df['DATA'].dt.to_period('M')
    return df


def calcular_projecao(df):
    meses_fat = sorted(df[df['FATURADO'] > 0]['MES'].unique())
    if not meses_fat:
        return [], 0.0

    ano = int(meses_fat[-1].year)

    # calcula media diaria de cada mes com dados
    medias_diarias = []
    for m in meses_fat:
        total_m  = float(df[df['MES'] == m]['FATURADO'].sum())
        dias_fat = max(int(df[(df['MES'] == m) & (df['FATURADO'] > 0)]['DATA'].nunique()), 1)
        medias_diarias.append(total_m / dias_fat)
    media_base = float(sum(medias_diarias) / len(medias_diarias))

    # dias uteis aproximados por mes (5/7 dos dias totais)
    rows = []
    acum = 0.0
    for idx in range(12):
        m_num    = idx + 1
        periodo  = pd.Period(f'{ano}-{m_num:02d}', 'M')
        fat_real = float(df[df['MES'] == periodo]['FATURADO'].sum())
        dias_c_fat = int(df[(df['MES'] == periodo) & (df['FATURADO'] > 0)]['DATA'].nunique())
        dias_tot   = DIAS_MES[idx]

        # dias uteis do mes (aproximacao: dias sem sabado/domingo)
        import calendar
        _, dias_no_mes = calendar.monthrange(ano, m_num)
        dias_uteis = sum(1 for d in range(1, dias_no_mes+1)
                         if datetime(ano, m_num, d).weekday() < 5)

        if fat_real > 0 and dias_c_fat >= (dias_uteis - 2):
            # mes completo (tolerancia de 2 dias)
            real   = fat_real
            proj   = fat_real
            status = 'Realizado'
        elif fat_real > 0:
            # mes parcial — extrapola pela media diaria do proprio mes
            media_mes = fat_real / max(dias_c_fat, 1)
            real      = fat_real
            proj      = media_mes * dias_uteis
            status    = 'Parcial'
        else:
            real   = None
            proj   = media_base * dias_uteis * FATOR_SAZ[idx]
            status = 'Projetado'

        acum += float(proj)
        rows.append({
            'mes':      MESES_PT[idx],
            'mes_abr':  MESES_ABR[idx],
            'num':      m_num,
            'dias':     dias_tot,
            'dias_uteis': dias_uteis,
            'real':     round(float(real), 2) if real is not None else None,
            'proj':     round(float(proj), 2),
            'acum':     round(acum, 2),
            'status':   status,
        })
    return rows, media_base


# ── ROTAS ──────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/arquivo-ativo')
def api_arquivo_ativo():
    return jsend({'nome': _excel_ativo['nome'], 'path': _excel_ativo['path']})


@app.route('/api/upload', methods=['POST'])
def api_upload():
    if 'arquivo' not in request.files:
        return jsend({'ok': False, 'erro': 'Nenhum arquivo enviado.'})
    f   = request.files['arquivo']
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ALLOWED_EXT:
        return jsend({'ok': False, 'erro': f'Formato nao suportado: {ext}. Use .xlsx ou .xls'})
    nome = secure_filename(f.filename)
    dest = os.path.join(UPLOAD_FOLDER, nome)
    f.save(dest)
    try:
        df_test = pd.read_excel(dest, nrows=5)
        if df_test.empty:
            raise ValueError('Planilha sem dados.')
    except Exception as e:
        os.remove(dest)
        return jsend({'ok': False, 'erro': str(e)})
    _excel_ativo['path'] = dest
    _excel_ativo['nome'] = nome
    return jsend({'ok': True, 'arquivo': nome})


@app.route('/api/resumo')
def api_resumo():
    try:
        df = load_data()
        rows, media_base = calcular_projecao(df)
        total_proj  = sum(r['proj'] for r in rows)
        total_real  = sum(r['real'] for r in rows if r['real'] is not None)
        media_m     = total_proj / 12 if rows else 0.0
        ultimo_fat  = df[df['FATURADO'] > 0]['DATA'].max()
        return jsend({
            'total_proj':       round(total_proj, 2),
            'total_real':       round(total_real, 2),
            'media_mensal':     round(media_m, 2),
            'media_diaria':     round(media_base, 2),
            'ultimo_realizado': ultimo_fat.strftime('%d/%m/%Y') if pd.notna(ultimo_fat) else '-',
            'meses_projetados': int(sum(1 for r in rows if r['status'] == 'Projetado')),
            'atualizado_em':    datetime.now().strftime('%d/%m/%Y %H:%M'),
            'arquivo':          _excel_ativo['nome'],
        })
    except Exception as e:
        return jsend({'erro': str(e)}, 500)


@app.route('/api/mensal')
def api_mensal():
    try:
        df = load_data()
        rows, _ = calcular_projecao(df)
        return jsend(rows)
    except Exception as e:
        return jsend({'erro': str(e)}, 500)


@app.route('/api/diario')
def api_diario():
    try:
        df  = load_data()
        mes = request.args.get('mes', 'todos')
        dff = df[df['FATURADO'] > 0].copy()
        if mes != 'todos':
            try: dff = dff[dff['MES'] == pd.Period(mes, 'M')]
            except: pass
        result = []; acum = 0.0
        for _, row in dff.iterrows():
            acum += float(row['FATURADO'])
            result.append({
                'data':      row['DATA'].strftime('%d/%m'),
                'data_full': row['DATA'].strftime('%d/%m/%Y'),
                'valor':     round(float(row['FATURADO']), 2),
                'acumulado': round(acum, 2),
                'dia_semana':['Seg','Ter','Qua','Qui','Sex','Sab','Dom'][row['DATA'].weekday()],
            })
        return jsend(result)
    except Exception as e:
        return jsend({'erro': str(e)}, 500)


@app.route('/api/trimestral')
def api_trimestral():
    try:
        df = load_data()
        rows, _ = calcular_projecao(df)
        return jsend([
            {'trim': 'T1 Jan-Mar', 'valor': round(sum(r['proj'] for r in rows[0:3]),  2)},
            {'trim': 'T2 Abr-Jun', 'valor': round(sum(r['proj'] for r in rows[3:6]),  2)},
            {'trim': 'T3 Jul-Set', 'valor': round(sum(r['proj'] for r in rows[6:9]),  2)},
            {'trim': 'T4 Out-Dez', 'valor': round(sum(r['proj'] for r in rows[9:12]), 2)},
        ])
    except Exception as e:
        return jsend({'erro': str(e)}, 500)


@app.route('/api/tabela')
def api_tabela():
    try:
        df = load_data()
        rows, _ = calcular_projecao(df)
        media_m = sum(r['proj'] for r in rows) / 12 if rows else 1.0
        result  = []
        for r in rows:
            vs = round((r['proj'] / media_m - 1) * 100, 1)
            result.append({
                **r,
                'real_fmt': br(r['real']) if r['real'] is not None else '-',
                'proj_fmt': br(r['proj']),
                'acum_fmt': br(r['acum']),
                'vs_media': vs,
            })
        return jsend(result)
    except Exception as e:
        return jsend({'erro': str(e)}, 500)


@app.route('/api/meses-disponiveis')
def api_meses():
    try:
        df    = load_data()
        meses = sorted(df[df['FATURADO'] > 0]['MES'].unique())
        return jsend([str(m) for m in meses])
    except Exception as e:
        return jsend([])


@app.route('/api/meta-mensal')
def api_meta_mensal():
    try:
        import calendar
        meta = request.args.get('meta', 0, type=float)

        df   = load_data()
        hoje = datetime.now()
        ano  = hoje.year
        mes  = hoje.month

        periodo_atual = pd.Period(f'{ano}-{mes:02d}', 'M')
        fat_hoje = float(df[
            (df['MES'] == periodo_atual) &
            (df['DATA'].dt.date <= hoje.date())
        ]['FATURADO'].sum())

        _, dias_no_mes     = calendar.monthrange(ano, mes)
        dias_corridos_rest = dias_no_mes - hoje.day
        dias_uteis_total   = sum(1 for d in range(1, dias_no_mes+1)
                                 if datetime(ano, mes, d).weekday() < 5)
        dias_uteis_pass    = sum(1 for d in range(1, hoje.day+1)
                                 if datetime(ano, mes, d).weekday() < 5)
        dias_uteis_rest    = sum(1 for d in range(hoje.day+1, dias_no_mes+1)
                                 if datetime(ano, mes, d).weekday() < 5)

        falta          = max(meta - fat_hoje, 0) if meta > 0 else 0
        pct_atingido   = round(fat_hoje / meta * 100, 1) if meta > 0 else 0
        necessario_dia = round(falta / dias_uteis_rest, 2) if dias_uteis_rest > 0 else 0
        media_dia_real = round(fat_hoje / max(dias_uteis_pass, 1), 2)
        projecao_fim   = round(fat_hoje + media_dia_real * dias_uteis_rest, 2)
        status_meta    = 'atingida'   if fat_hoje >= meta > 0     else \
                         'no_caminho' if projecao_fim >= meta > 0 else \
                         'risco'      if meta > 0                 else 'sem_meta'

        return jsend({
            'meta':               round(meta, 2),
            'fat_hoje':           round(fat_hoje, 2),
            'falta':              round(falta, 2),
            'pct_atingido':       pct_atingido,
            'dias_uteis_rest':    dias_uteis_rest,
            'dias_corridos_rest': dias_corridos_rest,
            'dias_uteis_total':   dias_uteis_total,
            'dias_uteis_pass':    dias_uteis_pass,
            'necessario_dia':     necessario_dia,
            'media_dia_real':     media_dia_real,
            'projecao_fim_mes':   projecao_fim,
            'status_meta':        status_meta,
            'mes_nome':           MESES_PT[mes - 1],
            'hoje':               hoje.strftime('%d/%m/%Y'),
        })
    except Exception as e:
        return jsend({'erro': str(e)}, 500)


@app.route('/api/exportar-excel')
def api_exportar_excel():
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter

        df   = load_data()
        rows, media_base = calcular_projecao(df)
        media_m = sum(r['proj'] for r in rows) / 12 if rows else 1.0

        wb = Workbook()

        # ── ABA 1: PROJECAO ──────────────────────────────────────────────────
        ws = wb.active
        ws.title = 'Projecao Mensal'

        C_NAVY  = '1E2761'
        C_BLUE  = '3A5BA0'
        C_GOLD  = 'F0C030'
        C_WHITE = 'FFFFFF'
        C_LIGHT = 'EEF2FA'
        C_REAL  = 'D6EAF8'
        C_PARC  = 'FEF9E7'
        C_PROJ  = 'EAF4F4'
        C_TOT   = 'D5E8D4'

        def side(c='C8C8C8', s='thin'):
            return Side(border_style=s, color=c)
        brd = Border(left=side(), right=side(), top=side(), bottom=side())

        # titulo
        ws.merge_cells('A1:I1')
        c = ws['A1']
        c.value     = 'RELATORIO DE PROJECAO DE FATURAMENTO 2026'
        c.font      = Font(bold=True, color=C_WHITE, size=14)
        c.fill      = PatternFill('solid', fgColor=C_NAVY)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 34

        ws.merge_cells('A2:I2')
        c2 = ws['A2']
        c2.value     = f'Exportado em {datetime.now().strftime("%d/%m/%Y %H:%M")}  |  Arquivo: {_excel_ativo["nome"]}  |  Media diaria base: {br(media_base)}'
        c2.font      = Font(italic=True, size=9, color='555555')
        c2.fill      = PatternFill('solid', fgColor='F2F4F7')
        c2.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 16

        # cabecalho tabela
        hdrs = ['Mes', 'Dias', 'Dias Uteis', 'Faturamento Real (R$)',
                'Projecao Mensal (R$)', 'Acumulado (R$)', 'vs Media (%)', 'Status', 'Fator Sazonal']
        for col_i, h in enumerate(hdrs, 1):
            c = ws.cell(row=4, column=col_i, value=h)
            c.font      = Font(bold=True, color=C_WHITE, size=10)
            c.fill      = PatternFill('solid', fgColor=C_BLUE)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border    = brd
        ws.row_dimensions[4].height = 28

        bg_map = {'Realizado': C_REAL, 'Parcial': C_PARC, 'Projetado': C_PROJ}
        FATOR_NOMES = [1.05,0.95,1.00,0.98,1.02,0.97,0.96,1.00,1.03,1.05,1.08,1.10]

        for i, r in enumerate(rows):
            row_n  = i + 5
            vs     = round((r['proj'] / media_m - 1) * 100, 1)
            bg     = bg_map.get(r['status'], C_LIGHT)
            vals   = [
                r['mes'], r['dias'], r['dias_uteis'],
                r['real'] if r['real'] is not None else '',
                r['proj'], r['acum'], vs / 100,
                r['status'], FATOR_NOMES[i]
            ]
            fmts = [None, None, None,
                    'R$ #,##0.00', 'R$ #,##0.00', 'R$ #,##0.00',
                    '+0.0%;-0.0%;0.0%', None, '0.00']
            for col_i, (val, fmt) in enumerate(zip(vals, fmts), 1):
                c = ws.cell(row=row_n, column=col_i, value=val)
                c.fill      = PatternFill('solid', fgColor=bg)
                c.border    = brd
                c.alignment = Alignment(horizontal='center' if col_i not in [1,8] else 'left',
                                        vertical='center')
                if fmt: c.number_format = fmt
                if col_i in [5, 6]: c.font = Font(bold=True, size=10)
            ws.row_dimensions[row_n].height = 20

        # total
        r_tot = len(rows) + 5
        total_real = sum(r['real'] for r in rows if r['real'] is not None)
        total_proj = sum(r['proj'] for r in rows)
        tot_vals   = ['TOTAL ANUAL', '', '', total_real, total_proj, total_proj, '', '', '']
        tot_fmts   = [None, None, None, 'R$ #,##0.00', 'R$ #,##0.00', 'R$ #,##0.00', None, None, None]
        for col_i, (val, fmt) in enumerate(zip(tot_vals, tot_fmts), 1):
            c = ws.cell(row=r_tot, column=col_i, value=val)
            c.font      = Font(bold=True, color=C_WHITE, size=10)
            c.fill      = PatternFill('solid', fgColor=C_NAVY)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = brd
            if fmt: c.number_format = fmt
        ws.row_dimensions[r_tot].height = 24

        # larguras
        widths = [14, 7, 10, 22, 22, 22, 13, 12, 13]
        for col_i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_i)].width = w

        # ── ABA 2: DADOS DIARIOS ─────────────────────────────────────────────
        ws2 = wb.create_sheet('Dados Diarios')
        ws2.merge_cells('A1:E1')
        ws2['A1'].value     = 'HISTORICO DIARIO DE FATURAMENTO'
        ws2['A1'].font      = Font(bold=True, color=C_WHITE, size=12)
        ws2['A1'].fill      = PatternFill('solid', fgColor=C_NAVY)
        ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws2.row_dimensions[1].height = 28

        hdrs2 = ['Data', 'Dia da Semana', 'Mes', 'Faturado (R$)', 'Acumulado (R$)']
        for col_i, h in enumerate(hdrs2, 1):
            c = ws2.cell(row=2, column=col_i, value=h)
            c.font      = Font(bold=True, color=C_WHITE, size=10)
            c.fill      = PatternFill('solid', fgColor=C_BLUE)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = brd

        dias_sem = ['Segunda','Terca','Quarta','Quinta','Sexta','Sabado','Domingo']
        df_exp   = df[df['FATURADO'] >= 0].sort_values('DATA').copy()
        acum2    = 0.0
        for idx2, (_, drow) in enumerate(df_exp.iterrows()):
            rn = idx2 + 3
            acum2 += float(drow['FATURADO'])
            bg_r = 'F4F7FF' if idx2 % 2 == 0 else 'FFFFFF'
            vals2 = [drow['DATA'].date(), dias_sem[drow['DATA'].weekday()],
                     drow['DATA'].strftime('%B/%Y'), float(drow['FATURADO']), acum2]
            fmts2 = ['DD/MM/YYYY', None, None, 'R$ #,##0.00', 'R$ #,##0.00']
            for col_i, (val, fmt) in enumerate(zip(vals2, fmts2), 1):
                c = ws2.cell(row=rn, column=col_i, value=val)
                c.fill      = PatternFill('solid', fgColor=bg_r)
                c.border    = brd
                c.alignment = Alignment(horizontal='center', vertical='center')
                if fmt: c.number_format = fmt
            ws2.row_dimensions[rn].height = 18

        for col_i, w in enumerate([14, 14, 16, 20, 20], 1):
            ws2.column_dimensions[get_column_letter(col_i)].width = w

        # ── SALVAR E RETORNAR ────────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        nome_arq = f'Projecao_Faturamento_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        return Response(
            buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{nome_arq}"'}
        )
    except Exception as e:
        import traceback
        return jsend({'erro': str(e), 'detalhe': traceback.format_exc()}, 500)


if __name__ == '__main__':
    import socket
    PORT = int(os.environ.get('PORT', 9090))
    HOST = '0.0.0.0'
    try:
        ip_local = socket.gethostbyname(socket.gethostname())
    except:
        ip_local = '127.0.0.1'
    print(f'\n  Dashboard iniciado!')
    print(f'  Local:  http://127.0.0.1:{PORT}')
    print(f'  Rede:   http://{ip_local}:{PORT}')
    print(f'  (Compartilhe o endereco de Rede com quem estiver no mesmo Wi-Fi)\n')
    serve(app, host=HOST, port=PORT, threads=4)
