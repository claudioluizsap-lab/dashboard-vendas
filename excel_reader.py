import openpyxl

def ler_credenciais(caminho_arquivo):
    wb = openpyxl.load_workbook(caminho_arquivo)
    ws = wb.active

    headers = [str(cell.value).strip().upper() if cell.value else "" for cell in ws[1]]

    try:
        idx_site = headers.index("SITE")
        idx_usuario = headers.index("USUARIO")
        idx_senha = headers.index("SENHA")
    except ValueError as e:
        raise ValueError(f"Coluna nao encontrada na planilha: {e}. Colunas encontradas: {headers}")

    credenciais = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        site = row[idx_site]
        usuario = row[idx_usuario]
        senha = row[idx_senha]
        if site:
            credenciais.append({
                "site": str(site).strip(),
                "usuario": str(usuario).strip() if usuario else "",
                "senha": str(senha).strip() if senha else ""
            })

    return credenciais
