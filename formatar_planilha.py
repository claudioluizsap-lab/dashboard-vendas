import msoffcrypto, io, openpyxl, pandas as pd, warnings
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import datetime

warnings.filterwarnings("ignore")

# --- Descriptografar ---
with open("2025_temp.xlsx", "rb") as f:
    office_file = msoffcrypto.OfficeFile(f)
    office_file.load_key(password="1515")
    decrypted = io.BytesIO()
    office_file.decrypt(decrypted)

decrypted.seek(0)
wb_orig = openpyxl.load_workbook(decrypted, data_only=True)

# ========== Estilos ==========
COR_HEADER     = "1F3864"   # azul escuro
COR_HEADER2    = "2E75B6"   # azul médio
COR_SUBTOTAL   = "BDD7EE"   # azul claro
COR_ALT        = "EBF3FB"   # linha alternada
COR_BRANCO     = "FFFFFF"
COR_VERDE      = "375623"
COR_AMARELO    = "FFC000"
COR_TEXTO_BR   = "FFFFFF"
COR_TEXTO_ESC  = "1F3864"

def fill(cor): return PatternFill("solid", fgColor=cor)
def fonte(bold=False, cor="000000", size=10, italic=False):
    return Font(bold=bold, color=cor, size=size, italic=italic, name="Calibri")
def borda_fina():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)
def alinhar(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def estilizar_header(cel, texto=None, subtitulo=False):
    if texto is not None:
        cel.value = texto
    cor = COR_HEADER2 if subtitulo else COR_HEADER
    cel.fill = fill(cor)
    cel.font = fonte(bold=True, cor=COR_TEXTO_BR, size=10)
    cel.alignment = alinhar("center")
    cel.border = borda_fina()

def formatar_moeda(cel, valor):
    cel.value = valor if isinstance(valor, (int, float)) else 0
    cel.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'
    cel.border = borda_fina()
    cel.alignment = alinhar("right")

MESES_PT = ["Jan","Fev","Mar","Abr","Mai","Jun",
            "Jul","Ago","Set","Out","Nov","Dez"]

# ============================================================
# ABA 1 - CALCULOS (reformatada como "Despesas por Categoria")
# ============================================================
ws_calc = wb_orig["CALCULOS "]

linhas_dados = []
categorias = []
for row in ws_calc.iter_rows(min_row=2, values_only=True):
    cat = row[0]
    if cat and str(cat).strip():
        categorias.append(str(cat).strip())
        linhas_dados.append(row)

# Cabeçalhos de data (linha 1 a partir da col 3)
header_row = next(ws_calc.iter_rows(min_row=1, max_row=1, values_only=True))
datas = [c for c in header_row[2:] if isinstance(c, datetime.datetime)]

# ============================================================
# ABA 2 - Bd_dadosR (banco de dados principal)
# ============================================================
ws_bd = wb_orig["Bd_dadosR"]
bd_rows = list(ws_bd.iter_rows(min_row=1, values_only=True))
bd_header = bd_rows[0]
bd_data   = bd_rows[1:]

# ============================================================
# Criar novo workbook profissional
# ============================================================
wb_new = Workbook()
wb_new.remove(wb_new.active)

# ---- Aba CAPA ----
ws_capa = wb_new.create_sheet("Capa")
ws_capa.sheet_view.showGridLines = False
ws_capa.column_dimensions["A"].width = 2
ws_capa.column_dimensions["B"].width = 60
ws_capa.column_dimensions["C"].width = 20

# Título
ws_capa.merge_cells("B2:C2")
c = ws_capa["B2"]
c.value = "RELATÓRIO FINANCEIRO 2025"
c.font = Font(bold=True, size=20, color=COR_TEXTO_BR, name="Calibri")
c.fill = fill(COR_HEADER)
c.alignment = alinhar("center")
ws_capa.row_dimensions[2].height = 40

ws_capa.merge_cells("B3:C3")
c = ws_capa["B3"]
c.value = "Controle de Compras, Despesas e Projeções"
c.font = Font(size=12, color=COR_TEXTO_BR, name="Calibri")
c.fill = fill(COR_HEADER2)
c.alignment = alinhar("center")
ws_capa.row_dimensions[3].height = 22

for r in range(4, 10):
    ws_capa.row_dimensions[r].height = 8

abas = [
    ("Despesas por Categoria", "Resumo mensal de despesas por conta (2021-2026)"),
    ("Banco de Dados",         "Dados detalhados por conta e subconta"),
    ("Compras",                "Controle de compras orçadas x realizadas"),
    ("Orçado x Realizado",     "Comparativo orçado x realizado"),
    ("Projeção Compras",       "Projeção sobre compras futuras"),
]
ws_capa["B10"].value = "CONTEÚDO DA PLANILHA"
ws_capa["B10"].font = Font(bold=True, size=11, color=COR_TEXTO_BR, name="Calibri")
ws_capa["B10"].fill = fill(COR_HEADER)
ws_capa["B10"].alignment = alinhar("center")
ws_capa["C10"].fill = fill(COR_HEADER)

for i, (nome, desc) in enumerate(abas, start=11):
    c1 = ws_capa.cell(row=i, column=2, value=nome)
    c2 = ws_capa.cell(row=i, column=3, value=desc)
    bg = COR_ALT if i % 2 == 0 else COR_BRANCO
    c1.fill = fill(bg); c2.fill = fill(bg)
    c1.font = Font(bold=True, size=10, color=COR_TEXTO_ESC, name="Calibri")
    c2.font = Font(size=10, color=COR_TEXTO_ESC, name="Calibri")
    c1.border = borda_fina(); c2.border = borda_fina()
    c1.alignment = alinhar("left"); c2.alignment = alinhar("left")
    ws_capa.row_dimensions[i].height = 18

ws_capa.cell(row=len(abas)+13, column=2,
    value=f"Gerado em: {datetime.date.today().strftime('%d/%m/%Y')}").font = \
    Font(italic=True, size=9, color="888888", name="Calibri")

# ---- Aba DESPESAS POR CATEGORIA ----
ws_d = wb_new.create_sheet("Despesas por Categoria")
ws_d.sheet_view.showGridLines = False
ws_d.freeze_panes = "C3"

# Título da aba
ws_d.merge_cells(f"A1:{get_column_letter(len(datas)+3)}1")
t = ws_d["A1"]
t.value = "DESPESAS POR CATEGORIA — 2021 a 2026"
t.font = Font(bold=True, size=13, color=COR_TEXTO_BR, name="Calibri")
t.fill = fill(COR_HEADER)
t.alignment = alinhar("center")
ws_d.row_dimensions[1].height = 28

# Montar cabeçalhos de datas
col_start = 3
estilizar_header(ws_d.cell(row=2, column=1), "Categoria")
estilizar_header(ws_d.cell(row=2, column=2), "Total Geral")
ws_d.column_dimensions["A"].width = 28
ws_d.column_dimensions["B"].width = 16

for idx, dt in enumerate(datas):
    col = col_start + idx
    label = f"{MESES_PT[dt.month-1]}/{str(dt.year)[2:]}"
    c = ws_d.cell(row=2, column=col)
    estilizar_header(c, label, subtitulo=True)
    ws_d.column_dimensions[get_column_letter(col)].width = 11
ws_d.row_dimensions[2].height = 20

# Preencher dados
for row_i, row in enumerate(linhas_dados, start=3):
    cat = str(row[0]).strip() if row[0] else ""
    valores = [row[2 + i] if (2 + i) < len(row) and isinstance(row[2 + i], (int, float)) else 0
               for i in range(len(datas))]
    total = sum(v for v in valores if v)

    bg = COR_ALT if row_i % 2 == 0 else COR_BRANCO

    c_cat = ws_d.cell(row=row_i, column=1, value=cat)
    c_cat.font = Font(size=10, name="Calibri", color=COR_TEXTO_ESC)
    c_cat.fill = fill(bg); c_cat.border = borda_fina()
    c_cat.alignment = alinhar("left")

    c_tot = ws_d.cell(row=row_i, column=2)
    formatar_moeda(c_tot, total)
    c_tot.fill = fill(COR_SUBTOTAL)
    c_tot.font = Font(bold=True, size=10, name="Calibri")

    for col_i, val in enumerate(valores):
        c = ws_d.cell(row=row_i, column=col_start + col_i)
        formatar_moeda(c, val)
        c.fill = fill(bg)
        c.font = Font(size=9, name="Calibri")

    ws_d.row_dimensions[row_i].height = 16

# Linha de totais
last_row = len(linhas_dados) + 3
ws_d.row_dimensions[last_row].height = 20
c_total_lbl = ws_d.cell(row=last_row, column=1, value="TOTAL GERAL")
c_total_lbl.font = Font(bold=True, size=10, color=COR_TEXTO_BR, name="Calibri")
c_total_lbl.fill = fill(COR_HEADER); c_total_lbl.border = borda_fina()
c_total_lbl.alignment = alinhar("center")

for col_i in range(len(datas) + 1):
    col = 2 + col_i
    if col == 2:
        total_col = sum(
            sum(row[2+i] for i in range(len(datas)) if (2+i) < len(row) and isinstance(row[2+i], (int,float)))
            for row in linhas_dados
        )
    else:
        idx = col_i - 1
        total_col = sum(
            row[2+idx] for row in linhas_dados
            if (2+idx) < len(row) and isinstance(row[2+idx], (int,float))
        )
    c = ws_d.cell(row=last_row, column=col)
    formatar_moeda(c, total_col)
    c.fill = fill(COR_HEADER)
    c.font = Font(bold=True, size=10, color=COR_TEXTO_BR, name="Calibri")
    c.border = borda_fina()

# ---- Aba BANCO DE DADOS ----
ws_b = wb_new.create_sheet("Banco de Dados")
ws_b.sheet_view.showGridLines = False
ws_b.freeze_panes = "C2"

# Filtrar colunas relevantes: Conta, SubConta + anos 2021-2025
datas_bd = []
cols_bd  = []
for i, h in enumerate(bd_header):
    if isinstance(h, datetime.datetime) and 2021 <= h.year <= 2025:
        datas_bd.append(h)
        cols_bd.append(i)

# Cabeçalho
ws_b.cell(row=1, column=1); estilizar_header(ws_b.cell(row=1, column=1), "Conta")
ws_b.cell(row=1, column=2); estilizar_header(ws_b.cell(row=1, column=2), "SubConta")
ws_b.column_dimensions["A"].width = 28
ws_b.column_dimensions["B"].width = 30
for j, dt in enumerate(datas_bd):
    col = 3 + j
    label = f"{MESES_PT[dt.month-1]}/{str(dt.year)[2:]}"
    estilizar_header(ws_b.cell(row=1, column=col), label, subtitulo=True)
    ws_b.column_dimensions[get_column_letter(col)].width = 10
ws_b.row_dimensions[1].height = 20

# Dados (máx 2000 linhas para não estourar)
count = 0
for r_i, row in enumerate(bd_data, start=2):
    conta  = str(row[0]).strip() if row[0] else ""
    sub    = str(row[1]).strip() if row[1] else ""
    if not conta and not sub:
        continue
    bg = COR_ALT if count % 2 == 0 else COR_BRANCO
    count += 1

    c1 = ws_b.cell(row=r_i, column=1, value=conta)
    c1.font = Font(size=9, name="Calibri", bold=bool(conta)); c1.fill = fill(bg)
    c1.border = borda_fina(); c1.alignment = alinhar("left")

    c2 = ws_b.cell(row=r_i, column=2, value=sub)
    c2.font = Font(size=9, name="Calibri"); c2.fill = fill(bg)
    c2.border = borda_fina(); c2.alignment = alinhar("left")

    for j, ci in enumerate(cols_bd):
        val = row[ci] if ci < len(row) and isinstance(row[ci], (int, float)) else 0
        c = ws_b.cell(row=r_i, column=3 + j)
        formatar_moeda(c, val)
        c.fill = fill(bg)
        c.font = Font(size=9, name="Calibri")
    ws_b.row_dimensions[r_i].height = 14

    if count >= 2000:
        break

print(f"Banco de Dados: {count} linhas exportadas")

# ---- Copiar abas operacionais simplificadas ----
def copiar_aba_simplificada(wb_src_sheet, wb_dest, nome_aba, titulo):
    ws = wb_dest.create_sheet(nome_aba)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Título
    max_c = wb_src_sheet.max_column
    ws.merge_cells(f"A1:{get_column_letter(max_c)}1")
    t = ws["A1"]
    t.value = titulo
    t.font = Font(bold=True, size=13, color=COR_TEXTO_BR, name="Calibri")
    t.fill = fill(COR_HEADER); t.alignment = alinhar("center")
    ws.row_dimensions[1].height = 26

    for r_i, row in enumerate(wb_src_sheet.iter_rows(values_only=True), start=2):
        for c_i, val in enumerate(row, start=1):
            c = ws.cell(row=r_i, column=c_i, value=val)
            if r_i == 2:
                c.fill = fill(COR_HEADER2)
                c.font = Font(bold=True, size=9, color=COR_TEXTO_BR, name="Calibri")
                c.alignment = alinhar("center")
                c.border = borda_fina()
            else:
                bg = COR_ALT if r_i % 2 == 0 else COR_BRANCO
                c.fill = fill(bg)
                c.font = Font(size=9, name="Calibri")
                c.border = borda_fina()
                if isinstance(val, (int, float)):
                    c.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'
                    c.alignment = alinhar("right")
                elif isinstance(val, datetime.datetime):
                    c.number_format = 'DD/MM/YYYY'
                    c.alignment = alinhar("center")
        ws.row_dimensions[r_i].height = 14

    for col in ws.columns:
        max_w = 10
        for cell in col:
            if cell.value:
                max_w = max(max_w, min(len(str(cell.value)) + 2, 25))
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_w

for src_name, dest_name, titulo in [
    (" COMPRAS",                  "Compras",          "COMPRAS — ORÇADO X REALIZADO"),
    (" Orçado x realizado  ",     "Orçado x Realizado","ORÇADO X REALIZADO"),
    ("PROJEÇÃO SOBRE COMPRAS",    "Projeção Compras",  "PROJEÇÃO SOBRE COMPRAS"),
]:
    if src_name in wb_orig.sheetnames:
        copiar_aba_simplificada(wb_orig[src_name], wb_new, dest_name, titulo)
        print(f"Aba '{dest_name}' copiada.")

# ---- Salvar ----
output_path = r"D:\Users\Claudio\OneDrive\Relatorio de vendas\2025_Profissional.xlsx"
wb_new.save(output_path)
print(f"\nArquivo salvo: {output_path}")
