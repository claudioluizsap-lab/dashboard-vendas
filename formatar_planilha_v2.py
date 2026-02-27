import msoffcrypto, io, openpyxl, warnings, datetime, copy
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

warnings.filterwarnings("ignore")

# ── descriptografar ──────────────────────────────────────────────────────────
with open("2025_temp.xlsx", "rb") as f:
    of = msoffcrypto.OfficeFile(f)
    of.load_key(password="1515")
    dec = io.BytesIO()
    of.decrypt(dec)

dec.seek(0)
wb_orig = openpyxl.load_workbook(dec, data_only=True)

# ── ler dados CALCULOS ────────────────────────────────────────────────────────
ws_calc = wb_orig["CALCULOS "]
header  = [c.value for c in list(ws_calc.iter_rows(min_row=1, max_row=1))[0]]
datas   = [c for c in header if isinstance(c, datetime.datetime)]

cats = []; raw_rows = []
for row in ws_calc.iter_rows(min_row=2, values_only=True):
    if row[0] and str(row[0]).strip():
        cats.append(str(row[0]).strip())
        raw_rows.append(row)

# montar DataFrame
idx_datas = [i for i, h in enumerate(header) if isinstance(h, datetime.datetime)]
data_dict = {"Categoria": cats}
for i, dt in enumerate(datas):
    col_label = dt.strftime("%Y-%m")
    data_dict[col_label] = [
        float(r[idx_datas[i]]) if idx_datas[i] < len(r) and isinstance(r[idx_datas[i]], (int, float)) else 0.0
        for r in raw_rows
    ]
df = pd.DataFrame(data_dict).set_index("Categoria")
# agrupar duplicados
df = df.groupby(df.index).sum()
# atualizar lista de categorias sem duplicados
cats = list(df.index)

# totais úteis
anos = sorted({dt.year for dt in datas})
MESES_PT = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]

def total_ano(ano):
    cols = [dt.strftime("%Y-%m") for dt in datas if dt.year == ano]
    return df[cols].sum(axis=1).sum() if cols else 0

def mensal_total(ano):
    cols = [dt.strftime("%Y-%m") for dt in datas if dt.year == ano]
    return df[cols].sum() if cols else pd.Series(dtype=float)

def top_cats(n=10):
    return df.sum(axis=1).nlargest(n)

# ── estilos ───────────────────────────────────────────────────────────────────
COR_NAVY   = "1A2B5C"
COR_AZUL   = "2563EB"
COR_AZUL2  = "3B82F6"
COR_CINZA  = "F1F5F9"
COR_CINZA2 = "E2E8F0"
COR_BRANCO = "FFFFFF"
COR_VERDE  = "16A34A"
COR_VERD2  = "DCFCE7"
COR_VERM   = "DC2626"
COR_VERM2  = "FEE2E2"
COR_AMAR   = "D97706"
COR_AMAR2  = "FEF3C7"
COR_ALT    = "EFF6FF"
COR_HEAD   = "DBEAFE"
COR_TEXTO  = "1E293B"

def F(cor): return PatternFill("solid", fgColor=cor)
def Fo(bold=False, cor=COR_TEXTO, size=10, italic=False, name="Calibri"):
    return Font(bold=bold, color=cor, size=size, italic=italic, name=name)
def B(c="CBD5E1", s="thin"):
    sd = Side(style=s, color=c)
    return Border(left=sd, right=sd, top=sd, bottom=sd)
def A(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def header_cell(ws, row, col, texto, cor_bg=None, cor_fg=COR_BRANCO, bold=True,
                size=10, h="center", border=True):
    c = ws.cell(row=row, column=col, value=texto)
    if cor_bg:
        c.fill = F(cor_bg)
    c.font = Fo(bold=bold, cor=cor_fg, size=size)
    c.alignment = A(h=h)
    if border:
        c.border = B()
    return c

def data_cell(ws, row, col, value, fmt=None, bg=None, bold=False,
              h="right", color=COR_TEXTO):
    c = ws.cell(row=row, column=col, value=value)
    if bg:
        c.fill = F(bg)
    c.font = Fo(bold=bold, cor=color, size=9)
    c.alignment = A(h=h)
    c.border = B()
    if fmt:
        c.number_format = fmt
    return c

FMT_BRL  = 'R$ #,##0.00;[Red]-R$ #,##0.00'
FMT_BRL0 = 'R$ #,##0;[Red]-R$ #,##0'
FMT_PCT  = '0.0%'
FMT_DATE = 'MM/AAAA'

# ── criar workbook ────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ═════════════════════════════════════════════════════════════════════════════
# ABA 1: DASHBOARD
# ═════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Dashboard")
ws.sheet_view.showGridLines = False
ws.sheet_view.showRowColHeaders = False

# larguras
for i, w in enumerate([1,22,16,16,16,16,16,3,22,16,16,16,16,16,1], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
for r in range(1, 60):
    ws.row_dimensions[r].height = 18

# ── cabeçalho ──
ws.merge_cells("B1:G1")
c = ws["B1"]
c.value = "RELATÓRIO FINANCEIRO — CONTROLE DE DESPESAS"
c.font  = Font(bold=True, size=18, color=COR_BRANCO, name="Calibri")
c.fill  = F(COR_NAVY)
c.alignment = A("center")
ws.row_dimensions[1].height = 40

ws.merge_cells("B2:G2")
c = ws["B2"]
c.value = f"Período: Jan/2021 a Out/2026  |  Atualizado em: {datetime.date.today().strftime('%d/%m/%Y')}"
c.font  = Font(italic=True, size=10, color="94A3B8", name="Calibri")
c.fill  = F("0F172A")
c.alignment = A("center")
ws.row_dimensions[2].height = 20

# ── KPI cards (linha 4-7) ──
ws.row_dimensions[3].height = 10
ws.row_dimensions[4].height = 22
ws.row_dimensions[5].height = 32
ws.row_dimensions[6].height = 20
ws.row_dimensions[7].height = 12

kpi_data = [
    ("B", "E", f"TOTAL GERAL (2021-2026)", df.sum().sum(), FMT_BRL0, COR_AZUL, COR_BRANCO),
    ("C", "F", f"TOTAL 2025",  total_ano(2025), FMT_BRL0, COR_VERDE, COR_BRANCO),
    ("D", "G", f"TOTAL 2024",  total_ano(2024), FMT_BRL0, COR_AMAR,  COR_BRANCO),
]

kpi_cols = [
    ("B", "C", "TOTAL GERAL\n(2021-2026)", df.sum().sum(),   COR_NAVY),
    ("D", "E", "TOTAL 2025",              total_ano(2025),   COR_AZUL),
    ("F", "G", "TOTAL 2024",              total_ano(2024),   COR_VERDE),
]

for (c1, c2, label, valor, cor) in kpi_cols:
    col1 = openpyxl.utils.column_index_from_string(c1)
    col2 = openpyxl.utils.column_index_from_string(c2)
    ws.merge_cells(f"{c1}4:{c2}4")
    ws.merge_cells(f"{c1}5:{c2}5")
    ws.merge_cells(f"{c1}6:{c2}6")
    hc = ws[f"{c1}4"]
    hc.value = label; hc.fill = F(cor)
    hc.font  = Font(bold=True, size=9, color="BFDBFE", name="Calibri")
    hc.alignment = A("center")
    vc = ws[f"{c1}5"]
    vc.value = valor; vc.number_format = FMT_BRL0
    vc.fill = F(cor)
    vc.font = Font(bold=True, size=18, color=COR_BRANCO, name="Calibri")
    vc.alignment = A("center")
    sc = ws[f"{c1}6"]
    sc.fill = F(cor); sc.alignment = A("center")

# KPIs secundários linha 9-11
ws.row_dimensions[8].height = 10
ws.row_dimensions[9].height = 18
ws.row_dimensions[10].height = 26
ws.row_dimensions[11].height = 10

# media mensal 2025
cols_2025 = [dt.strftime("%Y-%m") for dt in datas if dt.year == 2025]
media_2025 = total_ano(2025)/len(cols_2025) if cols_2025 else 0
# categorias ativas
n_cats = int((df.sum(axis=1) != 0).sum())
# crescimento 2024->2025
cresc = (total_ano(2025)/total_ano(2024)-1) if total_ano(2024) else 0

kpi2 = [
    ("B","C", "MÉDIA MENSAL 2025",    media_2025, FMT_BRL0),
    ("D","E", "CATEGORIAS ATIVAS",    n_cats,     "#,##0"),
    ("F","G", "CRESCIMENTO 2024→2025",cresc,      "+0.0%;-0.0%;0.0%"),
]
for (c1,c2,lbl,val,fmt) in kpi2:
    col1=openpyxl.utils.column_index_from_string(c1)
    ws.merge_cells(f"{c1}9:{c2}9"); ws.merge_cells(f"{c1}10:{c2}10")
    lc=ws[f"{c1}9"]; lc.value=lbl
    lc.fill=F(COR_CINZA); lc.font=Font(bold=True,size=8,color="64748B",name="Calibri")
    lc.alignment=A("center")
    vc=ws[f"{c1}10"]; vc.value=val; vc.number_format=fmt
    cor_val = COR_VERDE if (fmt.startswith("+") and val>=0) else COR_TEXTO
    vc.fill=F(COR_BRANCO)
    vc.font=Font(bold=True,size=14,color=cor_val,name="Calibri")
    vc.alignment=A("center")
    for r in [9,10]:
        for col in range(col1, openpyxl.utils.column_index_from_string(c2)+1):
            ws.cell(row=r,column=col).border=B("CBD5E1","thin")

# ── Top 10 Categorias (tabela B13:G23) ──
ws.row_dimensions[12].height = 8
ws.merge_cells("B13:G13")
tc=ws["B13"]; tc.value="TOP 10 MAIORES DESPESAS (Total 2021-2026)"
tc.fill=F(COR_AZUL); tc.font=Font(bold=True,size=10,color=COR_BRANCO,name="Calibri")
tc.alignment=A("center")

header_cell(ws,14,2,"Categoria",COR_HEAD,COR_TEXTO,size=9,h="left")
header_cell(ws,14,3,"Total (R$)",COR_HEAD,COR_TEXTO,size=9)
header_cell(ws,14,4,"2023",COR_HEAD,COR_TEXTO,size=9)
header_cell(ws,14,5,"2024",COR_HEAD,COR_TEXTO,size=9)
header_cell(ws,14,6,"2025",COR_HEAD,COR_TEXTO,size=9)
header_cell(ws,14,7,"% Total",COR_HEAD,COR_TEXTO,size=9)
ws.row_dimensions[14].height=18

top10 = top_cats(10)
total_geral = df.sum().sum()
for i, (cat, total) in enumerate(top10.items()):
    r = 15 + i
    ws.row_dimensions[r].height = 16
    bg = COR_ALT if i%2==0 else COR_BRANCO
    cols23 = [dt.strftime("%Y-%m") for dt in datas if dt.year==2023]
    cols24 = [dt.strftime("%Y-%m") for dt in datas if dt.year==2024]
    cols25 = [dt.strftime("%Y-%m") for dt in datas if dt.year==2025]
    valid23 = [c for c in cols23 if c in df.columns]
    valid24 = [c for c in cols24 if c in df.columns]
    valid25 = [c for c in cols25 if c in df.columns]
    t23 = float(df.loc[cat, valid23].values.sum()) if valid23 else 0
    t24 = float(df.loc[cat, valid24].values.sum()) if valid24 else 0
    t25 = float(df.loc[cat, valid25].values.sum()) if valid25 else 0
    data_cell(ws,r,2,cat,bg=bg,h="left",bold=(i<3))
    data_cell(ws,r,3,total,FMT_BRL0,bg=bg,bold=(i<3))
    data_cell(ws,r,4,t23,FMT_BRL0,bg=bg)
    data_cell(ws,r,5,t24,FMT_BRL0,bg=bg)
    data_cell(ws,r,6,t25,FMT_BRL0,bg=bg)
    data_cell(ws,r,7,total/total_geral if total_geral else 0,FMT_PCT,bg=bg)

# linha total
r=25
ws.row_dimensions[r].height=18
header_cell(ws,r,2,"TOTAL",COR_NAVY,size=9,h="left")
c23=[dt.strftime("%Y-%m") for dt in datas if dt.year==2023 and dt.strftime("%Y-%m") in df.columns]
c24=[dt.strftime("%Y-%m") for dt in datas if dt.year==2024 and dt.strftime("%Y-%m") in df.columns]
c25=[dt.strftime("%Y-%m") for dt in datas if dt.year==2025 and dt.strftime("%Y-%m") in df.columns]
tot23=float(df.loc[list(top10.index),c23].values.sum()) if c23 else 0
tot24=float(df.loc[list(top10.index),c24].values.sum()) if c24 else 0
tot25=float(df.loc[list(top10.index),c25].values.sum()) if c25 else 0
for col,val in [(3,total_geral),(4,tot23),(5,tot24),(6,tot25),(7,1.0)]:
    c=ws.cell(row=r,column=col,value=val)
    c.fill=F(COR_NAVY); c.font=Font(bold=True,size=9,color=COR_BRANCO,name="Calibri")
    c.alignment=A("center"); c.border=B()
    c.number_format = FMT_BRL0 if col<7 else FMT_PCT

# ── GRÁFICO DE BARRAS MENSAIS (direita do dashboard) ──
ws.row_dimensions[12].height = 8
ws.merge_cells("I13:N13")
gc=ws["I13"]; gc.value="EVOLUÇÃO MENSAL DE DESPESAS — 2024 e 2025"
gc.fill=F(COR_VERDE); gc.font=Font(bold=True,size=10,color=COR_BRANCO,name="Calibri")
gc.alignment=A("center")

# Tabela auxiliar para o gráfico
ws.row_dimensions[14].height=18
header_cell(ws,14,9,"Mês",COR_HEAD,COR_TEXTO,size=9,h="center")
header_cell(ws,14,10,"2023",COR_HEAD,COR_TEXTO,size=9)
header_cell(ws,14,11,"2024",COR_HEAD,COR_TEXTO,size=9)
header_cell(ws,14,12,"2025",COR_HEAD,COR_TEXTO,size=9)

graf_rows = []
for m in range(1,13):
    r=15+m-1
    ws.row_dimensions[r].height=16
    lbl = MESES_PT[m-1]
    t23=sum(float(df[dt.strftime("%Y-%m")].sum()) for dt in datas if dt.year==2023 and dt.month==m)
    t24=sum(float(df[dt.strftime("%Y-%m")].sum()) for dt in datas if dt.year==2024 and dt.month==m)
    t25=sum(float(df[dt.strftime("%Y-%m")].sum()) for dt in datas if dt.year==2025 and dt.month==m)
    bg = COR_ALT if m%2==0 else COR_BRANCO
    data_cell(ws,r,9,lbl,bg=bg,h="center")
    data_cell(ws,r,10,t23,FMT_BRL0,bg=bg)
    data_cell(ws,r,11,t24,FMT_BRL0,bg=bg)
    data_cell(ws,r,12,t25,FMT_BRL0,bg=bg)

# criar gráfico de linhas
chart = LineChart()
chart.title   = "Despesas Mensais 2023-2025"
chart.style   = 10
chart.y_axis.title = "R$"
chart.x_axis.title = "Mês"
chart.width   = 18
chart.height  = 12

cats_ref = Reference(ws, min_col=9, min_row=15, max_row=26)
for col, serie_nome in [(10,"2023"),(11,"2024"),(12,"2025")]:
    data_ref = Reference(ws, min_col=col, min_row=14, max_row=26)
    chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.series[0].graphicalProperties.line.solidFill = "94A3B8"
chart.series[1].graphicalProperties.line.solidFill = "2563EB"
chart.series[2].graphicalProperties.line.solidFill = "16A34A"
for s in chart.series:
    s.graphicalProperties.line.width = 20000
ws.add_chart(chart, "I27")

# ═════════════════════════════════════════════════════════════════════════════
# ABA 2: RESUMO ANUAL
# ═════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Resumo Anual")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "B3"

ws2.column_dimensions["A"].width = 32
for i,ano in enumerate(anos):
    ws2.column_dimensions[get_column_letter(2+i)].width = 16
ws2.column_dimensions[get_column_letter(2+len(anos))].width = 16
ws2.column_dimensions[get_column_letter(3+len(anos))].width = 12

# título
ws2.merge_cells(f"A1:{get_column_letter(3+len(anos))}1")
t=ws2["A1"]; t.value="RESUMO ANUAL DE DESPESAS POR CATEGORIA"
t.font=Font(bold=True,size=14,color=COR_BRANCO,name="Calibri")
t.fill=F(COR_NAVY); t.alignment=A("center"); ws2.row_dimensions[1].height=34

# cabeçalhos
header_cell(ws2,2,1,"Categoria",COR_HEAD,COR_TEXTO,size=10,h="left")
for i,ano in enumerate(anos):
    header_cell(ws2,2,2+i,str(ano),COR_HEAD,COR_TEXTO,size=10)
header_cell(ws2,2,2+len(anos),"TOTAL GERAL",COR_AZUL,size=10)
header_cell(ws2,2,3+len(anos),"% Total",COR_AZUL,size=10)
ws2.row_dimensions[2].height=22

# dados por categoria
total_geral = df.sum().sum()
for idx, cat in enumerate(cats):
    r = 3+idx
    ws2.row_dimensions[r].height=15
    bg = COR_ALT if idx%2==0 else COR_BRANCO
    total_cat = float(df.loc[cat].sum())

    c=ws2.cell(row=r,column=1,value=cat)
    c.fill=F(bg); c.font=Fo(size=9,bold=(total_cat>50000)); c.border=B()
    c.alignment=A("left")

    for j,ano in enumerate(anos):
        cols_ano = [dt.strftime("%Y-%m") for dt in datas if dt.year==ano]
        val = float(df.loc[cat,cols_ano].sum()) if cols_ano else 0
        dc=ws2.cell(row=r,column=2+j,value=val)
        dc.fill=F(bg); dc.border=B(); dc.alignment=A("right")
        dc.font=Fo(size=9); dc.number_format=FMT_BRL0

    # total e %
    tc=ws2.cell(row=r,column=2+len(anos),value=total_cat)
    tc.fill=F(COR_HEAD); tc.border=B(); tc.alignment=A("right")
    tc.font=Fo(bold=True,size=9); tc.number_format=FMT_BRL0

    pc=ws2.cell(row=r,column=3+len(anos),value=total_cat/total_geral if total_geral else 0)
    pc.fill=F(bg); pc.border=B(); pc.alignment=A("center")
    pc.font=Fo(size=9); pc.number_format=FMT_PCT

# linha totais
r_tot=3+len(cats)
ws2.row_dimensions[r_tot].height=22
header_cell(ws2,r_tot,1,"TOTAL GERAL",COR_NAVY,size=10,h="left")
for j,ano in enumerate(anos):
    cols_ano=[dt.strftime("%Y-%m") for dt in datas if dt.year==ano]
    val=float(df[cols_ano].sum().sum()) if cols_ano else 0
    c=ws2.cell(row=r_tot,column=2+j,value=val)
    c.fill=F(COR_NAVY); c.font=Font(bold=True,size=10,color=COR_BRANCO,name="Calibri")
    c.border=B(); c.alignment=A("right"); c.number_format=FMT_BRL0

c=ws2.cell(row=r_tot,column=2+len(anos),value=total_geral)
c.fill=F(COR_NAVY); c.font=Font(bold=True,size=10,color=COR_BRANCO,name="Calibri")
c.border=B(); c.alignment=A("right"); c.number_format=FMT_BRL0

c=ws2.cell(row=r_tot,column=3+len(anos),value=1.0)
c.fill=F(COR_NAVY); c.font=Font(bold=True,size=10,color=COR_BRANCO,name="Calibri")
c.border=B(); c.alignment=A("center"); c.number_format=FMT_PCT

# formatação condicional — escala de cor nas colunas de anos
for j in range(len(anos)):
    col_l=get_column_letter(2+j)
    r_ini=3; r_fim=3+len(cats)-1
    rng=f"{col_l}{r_ini}:{col_l}{r_fim}"
    ws2.conditional_formatting.add(rng, ColorScaleRule(
        start_type="min",start_color="FFFFFF",
        mid_type="percentile",mid_value=50,mid_color="BFDBFE",
        end_type="max",end_color="1D4ED8"
    ))

# ═════════════════════════════════════════════════════════════════════════════
# ABA 3: DESPESAS MENSAIS DETALHADAS
# ═════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Despesas Mensais")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = "C3"

ws3.column_dimensions["A"].width = 2
ws3.column_dimensions["B"].width = 30
ws3.column_dimensions["C"].width = 15

# filtrar apenas 2024-2026 para não ficar muito largo
datas_filtradas = [dt for dt in datas if dt.year >= 2024]
for i in range(len(datas_filtradas)):
    ws3.column_dimensions[get_column_letter(4+i)].width = 11

# título
max_col = 3+len(datas_filtradas)
ws3.merge_cells(f"B1:{get_column_letter(max_col)}1")
t=ws3["B1"]; t.value="DESPESAS MENSAIS DETALHADAS — 2024 a 2026"
t.font=Font(bold=True,size=14,color=COR_BRANCO,name="Calibri")
t.fill=F(COR_NAVY); t.alignment=A("center"); ws3.row_dimensions[1].height=34

# cabeçalhos
header_cell(ws3,2,2,"Categoria",COR_NAVY,size=10,h="left")
header_cell(ws3,2,3,"Total",COR_AZUL,size=10)
for i,dt in enumerate(datas_filtradas):
    lbl=f"{MESES_PT[dt.month-1]}/{str(dt.year)[2:]}"
    cor = COR_VERDE if dt.year==2025 else (COR_AMAR if dt.year==2024 else COR_AZUL2)
    header_cell(ws3,2,4+i,lbl,cor,size=9)
ws3.row_dimensions[2].height=22

# Agrupar por conta principal (primeira palavra)
grupos = {}
for cat in cats:
    g = cat.split()[0] if cat else "OUTROS"
    grupos.setdefault(g, []).append(cat)

row_num=3
for grupo, grupo_cats in sorted(grupos.items()):
    # linha de grupo
    ws3.merge_cells(f"B{row_num}:{get_column_letter(max_col)}{row_num}")
    gc=ws3[f"B{row_num}"]; gc.value=f"  {grupo}"
    gc.fill=F("1E3A5F"); gc.font=Font(bold=True,size=10,color="93C5FD",name="Calibri")
    gc.alignment=A("left"); ws3.row_dimensions[row_num].height=18
    row_num+=1

    for cat in grupo_cats:
        ws3.row_dimensions[row_num].height=15
        bg=COR_ALT if row_num%2==0 else COR_BRANCO
        total_cat=float(df.loc[cat].sum())

        c=ws3.cell(row=row_num,column=2,value=f"   {cat}")
        c.fill=F(bg); c.font=Fo(size=9); c.border=B(); c.alignment=A("left")

        tc=ws3.cell(row=row_num,column=3,value=total_cat)
        tc.fill=F(COR_HEAD); tc.font=Fo(bold=True,size=9); tc.border=B()
        tc.alignment=A("right"); tc.number_format=FMT_BRL0

        for i,dt in enumerate(datas_filtradas):
            col_lbl=dt.strftime("%Y-%m")
            val=float(df.loc[cat,col_lbl]) if col_lbl in df.columns else 0
            dc=ws3.cell(row=row_num,column=4+i,value=val if val else None)
            dc.fill=F(bg); dc.border=B(); dc.alignment=A("right")
            dc.font=Fo(size=9)
            if val:
                dc.number_format=FMT_BRL0
        row_num+=1

# formatação condicional data bars em total
ws3.conditional_formatting.add(
    f"C3:C{row_num}",
    DataBarRule(start_type="min",start_value=0,
                end_type="max",end_value=None,
                color="2563EB")
)

# ═════════════════════════════════════════════════════════════════════════════
# ABA 4: BANCO DE DADOS (tabela estruturada)
# ═════════════════════════════════════════════════════════════════════════════
ws_bd_orig = wb_orig["Bd_dadosR"]
bd_rows    = list(ws_bd_orig.iter_rows(min_row=1, values_only=True))
bd_header  = bd_rows[0]
bd_data    = bd_rows[1:]

ws4 = wb.create_sheet("Banco de Dados")
ws4.sheet_view.showGridLines = False
ws4.freeze_panes = "C2"

datas_bd  = [h for h in bd_header if isinstance(h, datetime.datetime) and 2023 <= h.year <= 2025]
idx_bd    = [i for i,h in enumerate(bd_header) if isinstance(h,datetime.datetime) and 2023<=h.year<=2025]

ws4.column_dimensions["A"].width = 28
ws4.column_dimensions["B"].width = 30
for i in range(len(datas_bd)):
    ws4.column_dimensions[get_column_letter(3+i)].width = 11

ws4.merge_cells(f"A1:{get_column_letter(2+len(datas_bd))}1")
t=ws4["A1"]; t.value="BANCO DE DADOS — 2023 a 2025"
t.font=Font(bold=True,size=14,color=COR_BRANCO,name="Calibri")
t.fill=F(COR_NAVY); t.alignment=A("center"); ws4.row_dimensions[1].height=34

header_cell(ws4,2,1,"Conta",COR_NAVY,size=10,h="left")
header_cell(ws4,2,2,"SubConta",COR_NAVY,size=10,h="left")
for i,dt in enumerate(datas_bd):
    lbl=f"{MESES_PT[dt.month-1]}/{str(dt.year)[2:]}"
    cor = COR_VERDE if dt.year==2025 else (COR_AZUL if dt.year==2024 else COR_HEAD[:-2]+"B2")
    header_cell(ws4,2,3+i,lbl,COR_HEAD,COR_TEXTO,size=9)
ws4.row_dimensions[2].height=20

count=0
for row in bd_data:
    conta=str(row[0]).strip() if row[0] else ""
    sub  =str(row[1]).strip() if row[1] else ""
    if not conta and not sub: continue
    r=3+count; count+=1
    bg=COR_ALT if count%2==0 else COR_BRANCO
    ws4.row_dimensions[r].height=14

    c=ws4.cell(row=r,column=1,value=conta)
    c.fill=F(bg); c.font=Fo(size=9,bold=bool(conta)); c.border=B(); c.alignment=A("left")
    c=ws4.cell(row=r,column=2,value=sub)
    c.fill=F(bg); c.font=Fo(size=9); c.border=B(); c.alignment=A("left")

    for j,ci in enumerate(idx_bd):
        val=row[ci] if ci<len(row) and isinstance(row[ci],(int,float)) else None
        dc=ws4.cell(row=r,column=3+j,value=val if val else None)
        dc.fill=F(bg); dc.border=B(); dc.alignment=A("right")
        dc.font=Fo(size=9)
        if val: dc.number_format=FMT_BRL0
    if count>=1500: break

# ═════════════════════════════════════════════════════════════════════════════
# ABA 5: CAPA
# ═════════════════════════════════════════════════════════════════════════════
ws_capa = wb.create_sheet("Capa", 0)
ws_capa.sheet_view.showGridLines = False
ws_capa.sheet_view.showRowColHeaders = False

for col in range(1,10):
    ws_capa.column_dimensions[get_column_letter(col)].width = 14
for r in range(1,30):
    ws_capa.row_dimensions[r].height = 22

# fundo (apenas cor de fundo em cada célula, sem merge global)
for rr in range(1,31):
    for cc in range(1,10):
        ws_capa.cell(row=rr,column=cc).fill=F("0F172A")

# título principal
ws_capa.merge_cells("B3:H4")
t=ws_capa["B3"]
t.value = "RELATÓRIO FINANCEIRO"
t.font  = Font(bold=True, size=28, color=COR_BRANCO, name="Calibri")
t.fill  = F("0F172A"); t.alignment = A("center")
ws_capa.row_dimensions[3].height = 46
ws_capa.row_dimensions[4].height = 46

ws_capa.merge_cells("B5:H5")
t2=ws_capa["B5"]
t2.value = "CONTROLE DE DESPESAS E PROJEÇÕES"
t2.font  = Font(size=14, color="93C5FD", name="Calibri")
t2.fill  = F("0F172A"); t2.alignment = A("center")
ws_capa.row_dimensions[5].height = 28

ws_capa.merge_cells("B6:H6")
t3=ws_capa["B6"]
t3.value = f"Jan/2021 — Out/2026  |  {len(cats)} categorias  |  {datetime.date.today().strftime('%d/%m/%Y')}"
t3.font  = Font(italic=True, size=10, color="64748B", name="Calibri")
t3.fill  = F("0F172A"); t3.alignment = A("center")

# separador
for col in range(2,9):
    c=ws_capa.cell(row=7,column=col)
    c.fill=F("2563EB")
ws_capa.row_dimensions[7].height = 4

# índice de abas
abas_info = [
    ("Dashboard",          "KPIs, Top 10 e gráfico de evolução mensal"),
    ("Resumo Anual",       "Total por categoria e ano (2021-2026)"),
    ("Despesas Mensais",   "Detalhamento mensal por categoria (2024-2026)"),
    ("Banco de Dados",     "Registros detalhados (Conta/SubConta, 2023-2025)"),
]

ws_capa.row_dimensions[9].height = 26
ws_capa.merge_cells("B9:D9"); ws_capa.merge_cells("E9:H9")
nc=ws_capa["B9"]; nc.value="ABA"; nc.fill=F(COR_AZUL)
nc.font=Font(bold=True,size=10,color=COR_BRANCO,name="Calibri"); nc.alignment=A("center")
dc=ws_capa["E9"]; dc.value="CONTEÚDO"; dc.fill=F(COR_AZUL)
dc.font=Font(bold=True,size=10,color=COR_BRANCO,name="Calibri"); dc.alignment=A("center")

for i,(nome,desc) in enumerate(abas_info):
    r=10+i
    ws_capa.row_dimensions[r].height=22
    ws_capa.merge_cells(f"B{r}:D{r}"); ws_capa.merge_cells(f"E{r}:H{r}")
    bg = "1E293B" if i%2==0 else "0F172A"
    nc=ws_capa[f"B{r}"]; nc.value=f"  {nome}"
    nc.fill=F(bg); nc.font=Font(bold=True,size=10,color="93C5FD",name="Calibri")
    nc.alignment=A("left"); nc.border=Border(bottom=Side(style="thin",color="1E3A5F"))
    dc=ws_capa[f"E{r}"]; dc.value=desc
    dc.fill=F(bg); dc.font=Font(size=10,color="94A3B8",name="Calibri")
    dc.alignment=A("left"); dc.border=Border(bottom=Side(style="thin",color="1E3A5F"))

# totais destaque
ws_capa.row_dimensions[15].height = 12
for i, (lbl, val, cor) in enumerate([
    ("Total Geral",  total_geral,       COR_AZUL),
    ("Total 2025",   total_ano(2025),   COR_VERDE),
    ("Total 2024",   total_ano(2024),   COR_AMAR),
]):
    col_s = 2+i*2; col_e = col_s+1
    ws_capa.row_dimensions[16].height = 18
    ws_capa.row_dimensions[17].height = 30
    ws_capa.merge_cells(f"{get_column_letter(col_s)}16:{get_column_letter(col_e)}16")
    ws_capa.merge_cells(f"{get_column_letter(col_s)}17:{get_column_letter(col_e)}17")
    lc=ws_capa.cell(row=16,column=col_s,value=lbl)
    lc.fill=F(cor); lc.font=Font(size=9,color="BFDBFE",bold=True,name="Calibri")
    lc.alignment=A("center")
    vc=ws_capa.cell(row=17,column=col_s,value=val)
    vc.fill=F(cor); vc.font=Font(bold=True,size=14,color=COR_BRANCO,name="Calibri")
    vc.alignment=A("center"); vc.number_format=FMT_BRL0

# ── salvar ────────────────────────────────────────────────────────────────────
output = r"D:\Users\Claudio\OneDrive\Relatorio de vendas\2025_Profissional.xlsx"
wb.save(output)
print(f"Arquivo salvo: {output}")
print(f"Abas: {[s.title for s in wb.worksheets]}")
