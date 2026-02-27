import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Credenciais"

header_fill = PatternFill("solid", fgColor="1A6FC4")
header_font = Font(bold=True, color="FFFFFF", size=12)
header_align = Alignment(horizontal="center", vertical="center")

border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

headers = ["SITE", "USUARIO", "SENHA"]
for col, titulo in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=titulo)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = border

ws.row_dimensions[1].height = 28

dados = [
    ["mercadolivre.com.br",  "seu_email@email.com", "sua_senha"],
    ["amazon.com.br",        "seu_email@email.com", "sua_senha"],
    ["americanas.com.br",    "seu_email@email.com", "sua_senha"],
    ["magazineluiza.com.br", "seu_email@email.com", "sua_senha"],
    ["casasbahia.com.br",    "seu_email@email.com", "sua_senha"],
    ["shopee.com.br",        "seu_email@email.com", "sua_senha"],
    ["kabum.com.br",         "seu_email@email.com", "sua_senha"],
    ["extra.com.br",         "seu_email@email.com", "sua_senha"],
    ["pontofrio.com.br",     "",                    ""],
    ["submarino.com.br",     "",                    ""],
]

alt_fill = PatternFill("solid", fgColor="EEF4FF")
data_font = Font(size=11)
data_align = Alignment(horizontal="left", vertical="center")

for i, linha in enumerate(dados, start=2):
    fill = alt_fill if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    for col, valor in enumerate(linha, start=1):
        cell = ws.cell(row=i, column=col, value=valor)
        cell.fill = fill
        cell.font = data_font
        cell.alignment = data_align
        cell.border = border
    ws.row_dimensions[i].height = 22

ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 32
ws.column_dimensions["C"].width = 20

ws.freeze_panes = "A2"

wb.save("planilha_modelo.xlsx")
print("Planilha modelo criada: planilha_modelo.xlsx")
