import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import os
from openpyxl.utils import column_index_from_string

def openpyxl_col(ref):
    col_str = ''.join(c for c in ref if c.isalpha())
    return column_index_from_string(col_str)

# ── CAMINHOS ──────────────────────────────────────────────────────────────────
path_in  = 'Bd_dadosprojecao.xlsx'
img_path = 'grafico_projecao.png'
path_out = 'Relatorio_Projecao_2026.xlsx'

# ── CARREGAR DADOS ─────────────────────────────────────────────────────────────
df = pd.read_excel(path_in, sheet_name='Planilha1', parse_dates=['DATA'])
df = df.dropna(subset=['DATA'])
df['MES'] = df['DATA'].dt.to_period('M')

jan_real    = df[df['MES'] == pd.Period('2026-01', 'M')]['FATURADO'].sum()
fev_parcial = df[df['MES'] == pd.Period('2026-02', 'M')]['FATURADO'].sum()
dias_fev_fat = df[(df['MES'] == pd.Period('2026-02', 'M')) & (df['FATURADO'] > 0)]['DATA'].nunique()

media_dia_fev  = fev_parcial / dias_fev_fat
media_dia_jan  = jan_real / 31
media_diaria   = (media_dia_jan + media_dia_fev) / 2

meses_pt = ['Janeiro','Fevereiro','Marco','Abril','Maio','Junho',
            'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
dias_mes  = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
fator     = [1.05, 0.95, 1.00, 0.98, 1.02, 0.97, 0.96, 1.00, 1.03, 1.05, 1.08, 1.10]

rows = []
for m in range(1, 13):
    if m == 1:
        real = jan_real;    proj = jan_real;             status = 'Realizado'
    elif m == 2:
        real = fev_parcial; proj = media_dia_fev * 28;  status = 'Parcial'
    else:
        real = None;        proj = media_diaria * dias_mes[m-1] * fator[m-1]; status = 'Projetado'
    rows.append({'Mes': meses_pt[m-1], 'M': m, 'Dias': dias_mes[m-1],
                 'Real': real, 'Proj': proj, 'Status': status})

proj_df      = pd.DataFrame(rows)
total_proj   = proj_df['Proj'].sum()
total_real   = proj_df['Real'].sum()
media_mensal = proj_df['Proj'].mean()

trim_labels = ['T1 (Jan-Mar)', 'T2 (Abr-Jun)', 'T3 (Jul-Set)', 'T4 (Out-Dez)']
trim_vals   = [proj_df['Proj'][:3].sum(), proj_df['Proj'][3:6].sum(),
               proj_df['Proj'][6:9].sum(), proj_df['Proj'][9:12].sum()]

# ── GRAFICOS ───────────────────────────────────────────────────────────────────
cores_map = {'Realizado': '#2E86AB', 'Parcial': '#F6AE2D', 'Projetado': '#A8DADC'}
cores     = [cores_map[s] for s in proj_df['Status']]

fig, axes = plt.subplots(2, 2, figsize=(16, 10))
fig.patch.set_facecolor('#F8F9FA')
fig.suptitle('Relatorio de Projecao de Faturamento - 2026',
             fontsize=16, fontweight='bold', y=0.99, color='#1A1A2E')

# 1) Barras mensais
ax1 = axes[0, 0]
ax1.set_facecolor('#FFFFFF')
bars = ax1.bar(proj_df['Mes'], proj_df['Proj'] / 1000, color=cores, edgecolor='#FFFFFF', linewidth=0.6)
ax1.set_title('Projecao Mensal de Faturamento', fontsize=11, fontweight='bold', color='#1A1A2E')
ax1.set_ylabel('R$ Mil', fontsize=9)
ax1.tick_params(axis='x', rotation=45, labelsize=8)
ax1.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'R${x:.0f}k'))
ax1.axhline(media_mensal / 1000, color='#E63946', linestyle='--', linewidth=1.4,
            label=f'Media: R${media_mensal/1000:.1f}k')
leg_patches = [mpatches.Patch(color=cores_map[k], label=k) for k in cores_map]
leg_patches.append(plt.Line2D([0], [0], color='#E63946', linestyle='--', label=f'Media: R${media_mensal/1000:.1f}k'))
ax1.legend(handles=leg_patches, fontsize=7.5, loc='upper left')
ax1.spines['top'].set_visible(False); ax1.spines['right'].set_visible(False)
for bar, val in zip(bars, proj_df['Proj']):
    ax1.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 4,
             f'R${val/1000:.0f}k', ha='center', va='bottom', fontsize=7.5, fontweight='bold', color='#1A1A2E')

# 2) Acumulado
ax2 = axes[0, 1]
ax2.set_facecolor('#FFFFFF')
acum = proj_df['Proj'].cumsum().values
ax2.plot(range(12), acum / 1000, marker='o', color='#E63946', linewidth=2.2, markersize=7, zorder=5)
ax2.fill_between(range(12), acum / 1000, alpha=0.12, color='#E63946')
ax2.set_title('Faturamento Acumulado 2026', fontsize=11, fontweight='bold', color='#1A1A2E')
ax2.set_ylabel('R$ Mil', fontsize=9)
ax2.set_xticks(range(12))
ax2.set_xticklabels(proj_df['Mes'], rotation=45, fontsize=8)
ax2.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'R${x:.0f}k'))
ax2.spines['top'].set_visible(False); ax2.spines['right'].set_visible(False)
for x, y in zip(range(12), acum / 1000):
    ax2.annotate(f'R${y:.0f}k', (x, y), textcoords='offset points', xytext=(0, 7),
                 ha='center', fontsize=7, color='#1A1A2E')

# 3) Pizza por trimestre
ax3 = axes[1, 0]
ax3.set_facecolor('#FFFFFF')
cores_trim = ['#2E86AB', '#F6AE2D', '#A8DADC', '#E63946']
wedges, texts, autotexts = ax3.pie(
    trim_vals, labels=trim_labels, autopct='%1.1f%%',
    colors=cores_trim, startangle=90, textprops={'fontsize': 9},
    wedgeprops={'edgecolor': 'white', 'linewidth': 2})
for at in autotexts:
    at.set_fontweight('bold')
ax3.set_title('Distribuicao Trimestral do Faturamento', fontsize=11, fontweight='bold', color='#1A1A2E')

# 4) Diario historico
ax4 = axes[1, 1]
ax4.set_facecolor('#FFFFFF')
df_real = df[(df['DATA'] <= pd.Timestamp('2026-02-16')) & (df['FATURADO'] > 0)].copy().reset_index(drop=True)
ax4.bar(range(len(df_real)), df_real['FATURADO'].values / 1000, color='#2E86AB', width=0.85, alpha=0.85)
mm7 = df_real['FATURADO'].rolling(7).mean().values
ax4.plot(range(len(df_real)), mm7 / 1000, color='#E63946', linewidth=1.8, label='Media movel 7 dias')
ax4.set_title('Faturamento Diario Realizado (Jan-Fev)', fontsize=11, fontweight='bold', color='#1A1A2E')
ax4.set_ylabel('R$ Mil', fontsize=9)
ax4.legend(fontsize=8)
ax4.set_xticks([])
ax4.spines['top'].set_visible(False); ax4.spines['right'].set_visible(False)

plt.tight_layout(rect=[0, 0, 1, 0.97])
fig.savefig(img_path, dpi=150, bbox_inches='tight', facecolor=fig.get_facecolor())
plt.close(fig)
print(f'Graficos salvos: {img_path}')

# ── EXCEL ──────────────────────────────────────────────────────────────────────
wb = Workbook()

# cores e estilos
C_AZUL_ESCURO  = '1A3A5C'
C_AZUL_MED     = '2E86AB'
C_CINZA_CLARO  = 'F2F4F7'
C_BRANCO       = 'FFFFFF'
C_REALIZADO    = 'D6EAF8'
C_PARCIAL      = 'FEF9E7'
C_PROJETADO    = 'EAF4F4'
C_TOTAL        = 'D5E8D4'

def side(c='C8C8C8', s='thin'):
    return Side(border_style=s, color=c)

brd_full = Border(left=side(), right=side(), top=side(), bottom=side())
brd_thick_bot = Border(left=side(), right=side(), top=side(), bottom=side('1A3A5C','medium'))

def hdr_style(ws, cell_ref, text, bg=C_AZUL_ESCURO, fg=C_BRANCO, sz=11, bold=True, center=True):
    c = ws[cell_ref]
    c.value = text
    c.font  = Font(bold=bold, color=fg, size=sz)
    c.fill  = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='center' if center else 'left', vertical='center', wrap_text=True)
    return c

def money_fmt(ws, cell_ref, value, bg=C_BRANCO, bold=False):
    c = ws[cell_ref]
    c.value  = value
    c.number_format = 'R$ #,##0.00'
    c.font   = Font(bold=bold, size=10)
    c.fill   = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = brd_full
    return c

# ── ABA 1: RESUMO EXECUTIVO ────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = 'Resumo Executivo'

# cabecalho principal
ws1.merge_cells('A1:G1')
c = ws1['A1']
c.value = 'RELATORIO DE PROJECAO DE FATURAMENTO - 2026'
c.font  = Font(bold=True, color=C_BRANCO, size=14)
c.fill  = PatternFill('solid', fgColor=C_AZUL_ESCURO)
c.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 36

ws1.merge_cells('A2:G2')
c2 = ws1['A2']
c2.value = f'Gerado em 19/02/2026  |  Base de dados: Bd_dadosprojecao.xlsx  |  Ultimo realizado: 16/02/2026'
c2.font  = Font(italic=True, size=9, color='666666')
c2.fill  = PatternFill('solid', fgColor='EEF2F7')
c2.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[2].height = 18

# KPIs
ws1.row_dimensions[4].height = 50
kpis = [
    ('A4', 'Total Realizado\n(Jan-Fev)', f'R$ {total_real:,.2f}', 'D6EAF8', C_AZUL_ESCURO),
    ('C4', 'Projecao Anual\n2026',       f'R$ {total_proj:,.2f}', 'D5F5E3', '1E8449'),
    ('E4', 'Media Mensal\nProjetada',    f'R$ {media_mensal:,.2f}', 'FEF9E7', '9A7D0A'),
    ('G4', 'Meses Projetados',           '10 meses', 'F5EEF8', '6C3483'),
]
for ref, label, val, bg, fg in kpis:
    ws1.merge_cells(f'{ref}:{get_column_letter(openpyxl_col(ref)+1)}{ref[1:]}')
    c = ws1[ref]
    c.value = f'{label}\n{val}'
    c.font  = Font(bold=True, size=10, color=fg)
    c.fill  = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = brd_full

# espaco
ws1.row_dimensions[5].height = 8

# cabecalho da tabela
hdr_cols = ['Mes', 'Dias', 'Faturamento Real (R$)', 'Projecao Mensal (R$)', 'Acumulado (R$)', 'vs Media (%)', 'Status']
for col_i, hdr in enumerate(hdr_cols, start=1):
    ref = f'{get_column_letter(col_i)}6'
    ws1[ref].value = hdr
    ws1[ref].font  = Font(bold=True, color=C_BRANCO, size=10)
    ws1[ref].fill  = PatternFill('solid', fgColor=C_AZUL_MED)
    ws1[ref].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws1[ref].border = brd_full
ws1.row_dimensions[6].height = 28

bg_map = {'Realizado': C_REALIZADO, 'Parcial': C_PARCIAL, 'Projetado': C_PROJETADO}
acum_val = 0
for i, row in proj_df.iterrows():
    r = i + 7
    acum_val += row['Proj']
    vs_media  = (row['Proj'] / media_mensal - 1) * 100
    bg = bg_map[row['Status']]

    ws1[f'A{r}'].value = row['Mes'];      ws1[f'A{r}'].alignment = Alignment(horizontal='left', vertical='center')
    ws1[f'B{r}'].value = row['Dias'];     ws1[f'B{r}'].alignment = Alignment(horizontal='center', vertical='center')
    real_val = row['Real'] if row['Real'] is not None and not np.isnan(row['Real']) else None
    ws1[f'C{r}'].value = real_val
    ws1[f'C{r}'].number_format = 'R$ #,##0.00' if real_val else '@'
    ws1[f'D{r}'].value = row['Proj'];     ws1[f'D{r}'].number_format = 'R$ #,##0.00'
    ws1[f'E{r}'].value = acum_val;        ws1[f'E{r}'].number_format = 'R$ #,##0.00'
    ws1[f'F{r}'].value = vs_media / 100;  ws1[f'F{r}'].number_format = '+0.0%;-0.0%;0.0%'
    ws1[f'G{r}'].value = row['Status']

    for col_l in ['A','B','C','D','E','F','G']:
        c = ws1[f'{col_l}{r}']
        c.fill   = PatternFill('solid', fgColor=bg)
        c.border = brd_full
        if col_l not in ['A','G']:
            c.alignment = Alignment(horizontal='right' if col_l in ['C','D','E','F'] else 'center', vertical='center')
    ws1.row_dimensions[r].height = 20

# total
r_tot = 7 + 12
for col_l, val, fmt in [
        ('A','TOTAL ANUAL PROJETADO','@'),('B','','@'),
        ('C', total_real, 'R$ #,##0.00'),('D', total_proj, 'R$ #,##0.00'),
        ('E', total_proj, 'R$ #,##0.00'),('F','','@'),('G','','@')]:
    c = ws1[f'{col_l}{r_tot}']
    c.value = val
    c.font  = Font(bold=True, size=10, color=C_BRANCO)
    c.fill  = PatternFill('solid', fgColor=C_AZUL_ESCURO)
    c.alignment = Alignment(horizontal='center' if col_l in ['A','B','G'] else 'right', vertical='center')
    c.border = brd_full
    if fmt != '@': c.number_format = fmt
ws1.row_dimensions[r_tot].height = 24

# legenda
r_leg = r_tot + 2
ws1.merge_cells(f'A{r_leg}:G{r_leg}')
ws1[f'A{r_leg}'].value = 'Legenda:  Realizado = dados reais confirmados  |  Parcial = mes em andamento (projecao baseada na media diaria)  |  Projetado = estimativa com fator sazonal'
ws1[f'A{r_leg}'].font  = Font(italic=True, size=8, color='555555')
ws1[f'A{r_leg}'].alignment = Alignment(horizontal='left', vertical='center')

# larguras
ws1.column_dimensions['A'].width = 14
ws1.column_dimensions['B'].width = 7
ws1.column_dimensions['C'].width = 20
ws1.column_dimensions['D'].width = 22
ws1.column_dimensions['E'].width = 22
ws1.column_dimensions['F'].width = 13
ws1.column_dimensions['G'].width = 12

# inserir grafico
img = XLImage(img_path)
img.width  = 900
img.height = 560
ws1.add_image(img, f'A{r_leg + 2}')

# ── ABA 2: DADOS DIARIOS ───────────────────────────────────────────────────────
ws2 = wb.create_sheet('Dados Diarios')
ws2.merge_cells('A1:D1')
ws2['A1'].value = 'HISTORICO DE FATURAMENTO DIARIO - DADOS REAIS'
ws2['A1'].font  = Font(bold=True, color=C_BRANCO, size=12)
ws2['A1'].fill  = PatternFill('solid', fgColor=C_AZUL_ESCURO)
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 30

for col_i, hdr in enumerate(['Data', 'Dia da Semana', 'Faturado (R$)', 'Acumulado (R$)'], start=1):
    c = ws2[f'{get_column_letter(col_i)}2']
    c.value = hdr
    c.font  = Font(bold=True, color=C_BRANCO, size=10)
    c.fill  = PatternFill('solid', fgColor=C_AZUL_MED)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd_full
ws2.row_dimensions[2].height = 22

dias_semana_pt = ['Segunda','Terca','Quarta','Quinta','Sexta','Sabado','Domingo']
df_export = df[df['FATURADO'] >= 0].sort_values('DATA').copy()
acum2 = 0
for i, (_, row_d) in enumerate(df_export.iterrows()):
    r = i + 3
    acum2 += row_d['FATURADO']
    dow = dias_semana_pt[row_d['DATA'].weekday()]
    bg_row = C_CINZA_CLARO if i % 2 == 0 else C_BRANCO

    ws2[f'A{r}'].value = row_d['DATA'].date(); ws2[f'A{r}'].number_format = 'DD/MM/YYYY'
    ws2[f'B{r}'].value = dow
    ws2[f'C{r}'].value = row_d['FATURADO'];    ws2[f'C{r}'].number_format = 'R$ #,##0.00'
    ws2[f'D{r}'].value = acum2;                ws2[f'D{r}'].number_format = 'R$ #,##0.00'

    for col_l in ['A','B','C','D']:
        c = ws2[f'{col_l}{r}']
        c.fill   = PatternFill('solid', fgColor=bg_row)
        c.border = brd_full
        c.alignment = Alignment(horizontal='center' if col_l in ['A','B'] else 'right', vertical='center')
    ws2.row_dimensions[r].height = 18

ws2.column_dimensions['A'].width = 14
ws2.column_dimensions['B'].width = 14
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 18

# ── ABA 3: METODOLOGIA ────────────────────────────────────────────────────────
ws3 = wb.create_sheet('Metodologia')
ws3.column_dimensions['A'].width = 90

ws3.merge_cells('A1:A1')
ws3['A1'].value = 'METODOLOGIA DE PROJECAO'
ws3['A1'].font  = Font(bold=True, color=C_BRANCO, size=13)
ws3['A1'].fill  = PatternFill('solid', fgColor=C_AZUL_ESCURO)
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 30

linhas_met = [
    ('', ''),
    ('METODO DE CALCULO:', ''),
    ('1. Janeiro 2026', f'Faturamento real: R$ {jan_real:,.2f} (31 dias)  |  Media diaria: R$ {media_dia_jan:,.2f}'),
    ('2. Fevereiro 2026', f'Parcialmente realizado: R$ {fev_parcial:,.2f} ({dias_fev_fat} dias)  |  Projecao para 28 dias: R$ {media_dia_fev*28:,.2f}'),
    ('3. Media diaria base', f'R$ {media_diaria:,.2f}/dia (media aritmetica entre Jan e Fev)'),
    ('4. Meses futuros', 'Projecao = Media diaria base x Dias do mes x Fator sazonal'),
    ('', ''),
    ('FATORES SAZONAIS APLICADOS:', ''),
    *[(f'   {meses_pt[m]}', f'Fator: {fator[m]:.2f}  =>  Efeito: {(fator[m]-1)*100:+.1f}%') for m in range(12)],
    ('', ''),
    ('PREMISSAS:', ''),
    ('- Tendencia', 'Crescimento linear baseado na media de Jan-Fev 2026'),
    ('- Sazonalidade', 'Fatores estimados com base em padrao tipico de varejo/servicos'),
    ('- Dias uteis', 'Considerados todos os dias do mes (ajuste por feriados nao aplicado)'),
    ('- Revisao', 'Recomenda-se revisar a projecao mensalmente conforme novos dados'),
]

for i, (col_a, col_b) in enumerate(linhas_met, start=2):
    ws3[f'A{i}'].value = f'  {col_a}  {col_b}'
    bold = col_a.endswith(':') or col_a == ''
    ws3[f'A{i}'].font  = Font(bold=(col_a.endswith(':')), size=10, color=C_AZUL_ESCURO if col_a.endswith(':') else '333333')
    ws3[f'A{i}'].alignment = Alignment(horizontal='left', vertical='center')
    ws3.row_dimensions[i].height = 18

wb.save(path_out)
print(f'Relatorio salvo: {path_out}')
