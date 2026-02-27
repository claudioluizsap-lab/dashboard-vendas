import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Ler dados do CSV
df = pd.read_csv('exemplo_estoque.csv')

# Ordenar por quantidade decrescente
df = df.sort_values('Quantidade', ascending=False).reset_index(drop=True)

# Calcular totais e percentuais
df['Percentual'] = (df['Quantidade'] / df['Quantidade'].sum()) * 100
df['Percentual Acumulado'] = df['Percentual'].cumsum()

# Classificar em ABC
def classificar_abc(percentual_acum):
    if percentual_acum <= 80:
        return 'A'
    elif percentual_acum <= 95:
        return 'B'
    else:
        return 'C'

df['Classe ABC'] = df['Percentual Acumulado'].apply(classificar_abc)

# Criar DataFrame final com colunas formatadas
df_relatorio = df[['Código', 'Produto', 'Quantidade', 'Percentual', 'Percentual Acumulado', 'Classe ABC']].copy()

# Salvar em Excel
arquivo_saida = 'Relatorio_ABC_Mercadorias.xlsx'
df_relatorio.to_excel(arquivo_saida, index=False, sheet_name='Análise ABC')

# Formatar o Excel
wb = load_workbook(arquivo_saida)
ws = wb['Análise ABC']

# Estilo do cabeçalho
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
header_alignment = Alignment(horizontal='center', vertical='center')

# Aplicar estilo no cabeçalho
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Bordas
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Formatar células de dados
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
    # Código
    row[0].alignment = Alignment(horizontal='center')
    row[0].border = thin_border
    
    # Produto
    row[1].alignment = Alignment(horizontal='left')
    row[1].border = thin_border
    
    # Quantidade
    row[2].alignment = Alignment(horizontal='center')
    row[2].number_format = '#,##0'
    row[2].border = thin_border
    
    # Percentual
    row[3].alignment = Alignment(horizontal='center')
    row[3].number_format = '0.00"%"'
    row[3].border = thin_border
    
    # Percentual Acumulado
    row[4].alignment = Alignment(horizontal='center')
    row[4].number_format = '0.00"%"'
    row[4].border = thin_border
    
    # Classe ABC
    classe = row[5].value
    row[5].alignment = Alignment(horizontal='center')
    row[5].font = Font(bold=True, size=11)
    row[5].border = thin_border
    
    # Colorir conforme a classe
    if classe == 'A':
        row[5].fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        row[5].font = Font(bold=True, size=11, color='FFFFFF')
    elif classe == 'B':
        row[5].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        row[5].font = Font(bold=True, size=11, color='FFFFFF')
    else:  # C
        row[5].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        row[5].font = Font(bold=True, size=11, color='FFFFFF')

# Ajustar largura das colunas
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 12

# Adicionar resumo
linha_resumo = ws.max_row + 2
ws[f'A{linha_resumo}'] = 'RESUMO DA ANÁLISE ABC'
ws[f'A{linha_resumo}'].font = Font(bold=True, size=12)
ws.merge_cells(f'A{linha_resumo}:F{linha_resumo}')
ws[f'A{linha_resumo}'].alignment = Alignment(horizontal='center')

# Contar itens por classe
resumo_classe = df_relatorio.groupby('Classe ABC').agg(
    Qtd_Itens=('Código', 'count'),
    Qtd_Total=('Quantidade', 'sum')
).reset_index()

linha_atual = linha_resumo + 2
ws[f'A{linha_atual}'] = 'Classe'
ws[f'B{linha_atual}'] = 'Qtd. Itens'
ws[f'C{linha_atual}'] = 'Qtd. Total'
ws[f'D{linha_atual}'] = '% do Total'

for cell in [ws[f'A{linha_atual}'], ws[f'B{linha_atual}'], ws[f'C{linha_atual}'], ws[f'D{linha_atual}']]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

linha_atual += 1
for _, row in resumo_classe.iterrows():
    ws[f'A{linha_atual}'] = row['Classe ABC']
    ws[f'B{linha_atual}'] = row['Qtd_Itens']
    ws[f'C{linha_atual}'] = row['Qtd_Total']
    ws[f'D{linha_atual}'] = (row['Qtd_Total'] / df['Quantidade'].sum()) * 100
    
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{linha_atual}'].border = thin_border
        ws[f'{col}{linha_atual}'].alignment = Alignment(horizontal='center')
    
    ws[f'C{linha_atual}'].number_format = '#,##0'
    ws[f'D{linha_atual}'].number_format = '0.00"%"'
    
    linha_atual += 1

wb.save(arquivo_saida)

print(f'Relatório ABC gerado com sucesso: {arquivo_saida}')
print(f'\nTotal de itens analisados: {len(df_relatorio)}')
print('\nDistribuição por Classe:')
print(resumo_classe.to_string(index=False))
