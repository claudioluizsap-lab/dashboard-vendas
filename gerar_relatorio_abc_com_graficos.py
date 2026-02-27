import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

# Ler dados do CSV
df = pd.read_csv('exemplo_estoque.csv')

# Ordenar por quantidade decrescente
df = df.sort_values('Quantidade', ascending=False).reset_index(drop=True)

# Calcular totais e percentuais
df['Percentual'] = (df['Quantidade'] / df['Quantidade'].sum()) * 100
df['Percentual Acumulado'] = df['Percentual'].cumsum()

# Classificar em ABC
def classificar_abc(percentual_acum):
    if percentual_acum <= 80:
        return 'A'
    elif percentual_acum <= 95:
        return 'B'
    else:
        return 'C'

df['Classe ABC'] = df['Percentual Acumulado'].apply(classificar_abc)

# Criar DataFrame final com colunas formatadas
df_relatorio = df[['Código', 'Produto', 'Quantidade', 'Percentual', 'Percentual Acumulado', 'Classe ABC']].copy()

# Salvar em Excel
arquivo_saida = 'Relatorio_ABC_Mercadorias.xlsx'
df_relatorio.to_excel(arquivo_saida, index=False, sheet_name='Análise ABC')

# Formatar o Excel
wb = load_workbook(arquivo_saida)
ws = wb['Análise ABC']

# Estilo do cabeçalho
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
header_alignment = Alignment(horizontal='center', vertical='center')

# Aplicar estilo no cabeçalho
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Bordas
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Formatar células de dados
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
    # Código
    row[0].alignment = Alignment(horizontal='center')
    row[0].border = thin_border
    
    # Produto
    row[1].alignment = Alignment(horizontal='left')
    row[1].border = thin_border
    
    # Quantidade
    row[2].alignment = Alignment(horizontal='center')
    row[2].number_format = '#,##0'
    row[2].border = thin_border
    
    # Percentual
    row[3].alignment = Alignment(horizontal='center')
    row[3].number_format = '0.00"%"'
    row[3].border = thin_border
    
    # Percentual Acumulado
    row[4].alignment = Alignment(horizontal='center')
    row[4].number_format = '0.00"%"'
    row[4].border = thin_border
    
    # Classe ABC
    classe = row[5].value
    row[5].alignment = Alignment(horizontal='center')
    row[5].font = Font(bold=True, size=11)
    row[5].border = thin_border
    
    # Colorir conforme a classe
    if classe == 'A':
        row[5].fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        row[5].font = Font(bold=True, size=11, color='FFFFFF')
    elif classe == 'B':
        row[5].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        row[5].font = Font(bold=True, size=11, color='FFFFFF')
    else:  # C
        row[5].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        row[5].font = Font(bold=True, size=11, color='FFFFFF')

# Ajustar largura das colunas
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 12

# Adicionar resumo
linha_resumo = ws.max_row + 2
ws[f'A{linha_resumo}'] = 'RESUMO DA ANÁLISE ABC'
ws[f'A{linha_resumo}'].font = Font(bold=True, size=12)
ws.merge_cells(f'A{linha_resumo}:F{linha_resumo}')
ws[f'A{linha_resumo}'].alignment = Alignment(horizontal='center')

# Contar itens por classe
resumo_classe = df_relatorio.groupby('Classe ABC').agg(
    Qtd_Itens=('Código', 'count'),
    Qtd_Total=('Quantidade', 'sum')
).reset_index()

linha_atual = linha_resumo + 2
linha_inicio_resumo = linha_atual
ws[f'A{linha_atual}'] = 'Classe'
ws[f'B{linha_atual}'] = 'Qtd. Itens'
ws[f'C{linha_atual}'] = 'Qtd. Total'
ws[f'D{linha_atual}'] = '% do Total'

for cell in [ws[f'A{linha_atual}'], ws[f'B{linha_atual}'], ws[f'C{linha_atual}'], ws[f'D{linha_atual}']]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

linha_atual += 1
for _, row in resumo_classe.iterrows():
    ws[f'A{linha_atual}'] = row['Classe ABC']
    ws[f'B{linha_atual}'] = row['Qtd_Itens']
    ws[f'C{linha_atual}'] = row['Qtd_Total']
    ws[f'D{linha_atual}'] = (row['Qtd_Total'] / df['Quantidade'].sum()) * 100
    
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{linha_atual}'].border = thin_border
        ws[f'{col}{linha_atual}'].alignment = Alignment(horizontal='center')
    
    ws[f'C{linha_atual}'].number_format = '#,##0'
    ws[f'D{linha_atual}'].number_format = '0.00"%"'
    
    linha_atual += 1

# ==================== ADICIONAR GRÁFICOS ====================

# 1. Gráfico de Barras - Top 10 Produtos
chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Top 10 Produtos por Quantidade"
chart1.y_axis.title = 'Quantidade'
chart1.x_axis.title = 'Produtos'

# Pegar apenas os top 10
max_produtos = min(10, len(df_relatorio))
data = Reference(ws, min_col=3, min_row=1, max_row=max_produtos+1)
cats = Reference(ws, min_col=2, min_row=2, max_row=max_produtos+1)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.height = 12
chart1.width = 20
chart1.legend = None

ws.add_chart(chart1, f"H2")

# 2. Gráfico de Pizza - Distribuição por Classe ABC
chart2 = PieChart()
chart2.title = "Distribuição de Quantidade por Classe ABC"
chart2.height = 12
chart2.width = 15

# Dados do resumo
labels = Reference(ws, min_col=1, min_row=linha_inicio_resumo+1, max_row=linha_inicio_resumo+3)
data = Reference(ws, min_col=3, min_row=linha_inicio_resumo, max_row=linha_inicio_resumo+3)
chart2.add_data(data, titles_from_data=True)
chart2.set_categories(labels)

# Adicionar labels com percentuais
chart2.dataLabels = DataLabelList()
chart2.dataLabels.showCatName = True
chart2.dataLabels.showPercent = True

ws.add_chart(chart2, f"H22")

# 3. Gráfico de Linha - Curva ABC (Percentual Acumulado)
from openpyxl.chart import LineChart

chart3 = LineChart()
chart3.title = "Curva ABC - Percentual Acumulado"
chart3.style = 13
chart3.y_axis.title = 'Percentual Acumulado (%)'
chart3.x_axis.title = 'Produtos (ordenados por quantidade)'

data = Reference(ws, min_col=5, min_row=1, max_row=len(df_relatorio)+1)
chart3.add_data(data, titles_from_data=True)
chart3.height = 12
chart3.width = 20

ws.add_chart(chart3, f"H42")

wb.save(arquivo_saida)

# ==================== GRÁFICOS PNG COM MATPLOTLIB ====================

# Criar diretório para gráficos se não existir
import os
if not os.path.exists('graficos'):
    os.makedirs('graficos')

# 1. Gráfico de Barras Horizontal - Top 15 produtos
plt.figure(figsize=(12, 8))
top15 = df_relatorio.head(15)
plt.barh(range(len(top15)), top15['Quantidade'], color='#4472C4')
plt.yticks(range(len(top15)), top15['Produto'], fontsize=9)
plt.xlabel('Quantidade', fontsize=11, fontweight='bold')
plt.title('Top 15 Produtos por Quantidade', fontsize=14, fontweight='bold', pad=20)
plt.gca().invert_yaxis()
for i, v in enumerate(top15['Quantidade']):
    plt.text(v + 2, i, str(v), va='center', fontsize=9)
plt.tight_layout()
plt.savefig('graficos/top15_produtos.png', dpi=150, bbox_inches='tight')
plt.close()

# 2. Gráfico de Pizza - Distribuição por Classe
fig, ax = plt.subplots(figsize=(10, 8))
colors_abc = {'A': '#00B050', 'B': '#FFC000', 'C': '#FF0000'}
colors = [colors_abc[c] for c in resumo_classe['Classe ABC']]
wedges, texts, autotexts = ax.pie(
    resumo_classe['Qtd_Total'], 
    labels=resumo_classe['Classe ABC'],
    autopct='%1.1f%%',
    colors=colors,
    startangle=90,
    textprops={'fontsize': 12, 'fontweight': 'bold', 'color': 'white'}
)
for autotext in autotexts:
    autotext.set_color('white')
    autotext.set_fontweight('bold')
ax.set_title('Distribuição de Quantidade por Classe ABC', fontsize=14, fontweight='bold', pad=20)

# Adicionar legenda
legend_labels = [f"Classe {row['Classe ABC']}: {row['Qtd_Itens']} itens ({row['Qtd_Total']} unidades)" 
                 for _, row in resumo_classe.iterrows()]
ax.legend(legend_labels, loc='upper left', bbox_to_anchor=(1, 0, 0.5, 1), fontsize=10)
plt.tight_layout()
plt.savefig('graficos/distribuicao_classe_abc.png', dpi=150, bbox_inches='tight')
plt.close()

# 3. Curva ABC - Linha do percentual acumulado
fig, ax = plt.subplots(figsize=(14, 7))
x = range(1, len(df_relatorio) + 1)
y = df_relatorio['Percentual Acumulado'].values

ax.plot(x, y, marker='o', linewidth=2, markersize=6, color='#4472C4')
ax.axhline(y=80, color='#00B050', linestyle='--', linewidth=1.5, label='Classe A (80%)')
ax.axhline(y=95, color='#FFC000', linestyle='--', linewidth=1.5, label='Classe B (95%)')
ax.fill_between(x, 0, y, alpha=0.2, color='#4472C4')

ax.set_xlabel('Produtos (ordenados por quantidade decrescente)', fontsize=11, fontweight='bold')
ax.set_ylabel('Percentual Acumulado (%)', fontsize=11, fontweight='bold')
ax.set_title('Curva ABC - Análise de Pareto', fontsize=14, fontweight='bold', pad=20)
ax.grid(True, alpha=0.3)
ax.legend(fontsize=10)
ax.set_ylim(0, 105)

plt.tight_layout()
plt.savefig('graficos/curva_abc.png', dpi=150, bbox_inches='tight')
plt.close()

# 4. Gráfico Combinado - Quantidade + Percentual Acumulado
fig, ax1 = plt.subplots(figsize=(14, 7))

color1 = '#4472C4'
ax1.set_xlabel('Produtos', fontsize=11, fontweight='bold')
ax1.set_ylabel('Quantidade', color=color1, fontsize=11, fontweight='bold')
bars = ax1.bar(range(len(df_relatorio)), df_relatorio['Quantidade'], color=color1, alpha=0.7, label='Quantidade')
ax1.tick_params(axis='y', labelcolor=color1)
ax1.set_xticks(range(len(df_relatorio)))
ax1.set_xticklabels(df_relatorio['Código'], rotation=45, ha='right', fontsize=8)

ax2 = ax1.twinx()
color2 = '#FF6B35'
ax2.set_ylabel('Percentual Acumulado (%)', color=color2, fontsize=11, fontweight='bold')
line = ax2.plot(range(len(df_relatorio)), df_relatorio['Percentual Acumulado'], 
                color=color2, marker='o', linewidth=2, markersize=5, label='% Acumulado')
ax2.tick_params(axis='y', labelcolor=color2)
ax2.axhline(y=80, color='#00B050', linestyle='--', linewidth=1, alpha=0.7)
ax2.axhline(y=95, color='#FFC000', linestyle='--', linewidth=1, alpha=0.7)
ax2.set_ylim(0, 105)

plt.title('Análise ABC - Quantidade e Percentual Acumulado', fontsize=14, fontweight='bold', pad=20)
fig.tight_layout()
plt.savefig('graficos/analise_combinada.png', dpi=150, bbox_inches='tight')
plt.close()

print(f'✓ Relatório ABC gerado com sucesso: {arquivo_saida}')
print(f'✓ Gráficos salvos na pasta: graficos/')
print(f'\nTotal de itens analisados: {len(df_relatorio)}')
print('\nDistribuição por Classe:')
print(resumo_classe.to_string(index=False))
print('\n📊 Gráficos gerados:')
print('  1. top15_produtos.png - Ranking dos 15 produtos com maior quantidade')
print('  2. distribuicao_classe_abc.png - Pizza mostrando % por classe')
print('  3. curva_abc.png - Curva de Pareto com percentual acumulado')
print('  4. analise_combinada.png - Gráfico duplo (barras + linha)')
